"""
Экспорт финансово-экономической модели ReFlex в формате БП МСП
(шаблон конкурса «Студенческий стартап» / МСП Банк)

Структура файла:
  Лист 1: Титульный лист     — информация о проекте (по шаблону МСП)
  Лист 2: БДДС               — бюджет движения денежных средств (по шаблону МСП)
  Лист 3: Инвест             — инвестиционные показатели: NPV, DCF, PP, PI (по шаблону МСП)
  Лист 4: График ДП          — график денежных потоков (График 1)
  Лист 5: Точка безубыточности — анализ breakeven + диаграмма
  Лист 6: Когорты пациентов  — когортная диаграмма
  Лист 7: Клиники            — таблица пользователей в клинике + подключение клиник
  Лист 8: Бизнес-логика      — автогенерированное описание бизнес-модели и стратегии
"""

from __future__ import annotations

import math
from datetime import date
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# Константы стилей
# ──────────────────────────────────────────────

_BLUE = "1F4E79"
_LIGHT_BLUE = "D6E4F0"
_GREEN_FILL = "E2EFDA"
_ORANGE_FILL = "FCE4D6"
_YELLOW_FILL = "FFF2CC"
_GRAY_FILL = "F2F2F2"
_WHITE = "FFFFFF"
_RED = "C00000"
_DARK_GREEN = "375623"

_HDR_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HDR_FONT = Font(bold=True, color=_WHITE, size=11)
_TITLE_FONT = Font(bold=True, size=13, color=_BLUE)
_BOLD_FONT = Font(bold=True, size=10)
_NORMAL_FONT = Font(size=10)
_ITALIC_FONT = Font(italic=True, size=10, color="595959")
_SECTION_FILL = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
_SECTION_FONT = Font(bold=True, size=10, color=_BLUE)
_GREEN_FILL_P = PatternFill(start_color=_GREEN_FILL, end_color=_GREEN_FILL, fill_type="solid")
_ORANGE_FILL_P = PatternFill(start_color=_ORANGE_FILL, end_color=_ORANGE_FILL, fill_type="solid")
_YELLOW_FILL_P = PatternFill(start_color=_YELLOW_FILL, end_color=_YELLOW_FILL, fill_type="solid")
_GRAY_FILL_P = PatternFill(start_color=_GRAY_FILL, end_color=_GRAY_FILL, fill_type="solid")
_BLUE_FILL_P = PatternFill(start_color=_LIGHT_BLUE, end_color=_LIGHT_BLUE, fill_type="solid")

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_MEDIUM_BORDER = Border(
    left=Side(style="medium"),
    right=Side(style="medium"),
    top=Side(style="medium"),
    bottom=Side(style="medium"),
)

_NUM_FMT = "#,##0"
_NUM_FMT_DEC = "#,##0.00"
_PCT_FMT = "0.0%"

# CAPM параметры (стандартные для МСП)
_RISK_FREE = 0.085
_BETA = 0.87
_MARKET_RETURN = 0.1411
_CAPM_RATE = _RISK_FREE + _BETA * (_MARKET_RETURN - _RISK_FREE)  # ≈ 13.38%

# Налоговая ставка УСН (доходы)
_TAX_RATE = 0.06

# Конвертация ₽ → тыс. ₽
_K = 1_000.0


# ──────────────────────────────────────────────
# Вспомогательные функции
# ──────────────────────────────────────────────

def _w(ws, row: int, col: int, value=None, *, font=None, fill=None, fmt=None,
        align: str = "left", bold: bool = False, border: bool = False,
        size: int = 10, color: str = "000000", italic: bool = False,
        wrap: bool = False) -> Any:
    """Запись в ячейку с применением стилей."""
    cell = ws.cell(row=row, column=col, value=value)
    f = Font(bold=bold, size=size, color=color, italic=italic)
    if font:
        f = font
    cell.font = f
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    cell.alignment = Alignment(
        horizontal=align,
        vertical="center",
        wrap_text=wrap,
    )
    if border:
        cell.border = _THIN_BORDER
    return cell


def _merge(ws, r1, c1, r2, c2, value=None, **kw):
    ws.merge_cells(
        start_row=r1, start_column=c1, end_row=r2, end_column=c2
    )
    return _w(ws, r1, c1, value, **kw)


def _rub(v: float) -> float:
    """Конвертируем ₽ → тыс. ₽ с округлением до 2 знаков."""
    return round(v / _K, 2)


def _safe(v, default=0.0):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return default
    return v


def _model_name(model_type: str) -> str:
    return {
        "model_a": "Модель A — B2B (клиника платит подписку)",
        "model_b": "Модель B — B2B2C (пациент платит аренду)",
        "model_ab": "Модель A+B — Гибридная",
    }.get(model_type, model_type)


def _sum_bank_by_month(bank_allocation: List[Dict], month_idx: int) -> float:
    """Возвращает total bank_used для месяца (0-indexed)."""
    if not bank_allocation or month_idx >= len(bank_allocation):
        return 0.0
    return _safe(bank_allocation[month_idx].get("bank_used", 0.0))


def _bank_category_totals(bank_allocation: List[Dict]) -> Dict[str, float]:
    """Агрегирует расходы банка (bank) по категориям за весь период через line_items."""
    totals: Dict[str, float] = {}
    for entry in (bank_allocation or []):
        line_items = entry.get("line_items", {})
        if line_items:
            for cat, item in line_items.items():
                if isinstance(item, dict):
                    bank_val = _safe(item.get("bank", 0.0))
                elif isinstance(item, (int, float)):
                    bank_val = _safe(item)
                else:
                    bank_val = 0.0
                totals[cat] = totals.get(cat, 0.0) + bank_val
    return totals


def _fixed_cost_labels() -> Dict[str, str]:
    return {
        "team_salaries": "Зарплаты команды",
        "infrastructure_fixed": "Инфраструктура (постоянная)",
        "office_rent": "Аренда офиса",
        "legal_services": "Юридические услуги",
        "other_fixed": "Прочие постоянные затраты",
    }


def _variable_cost_labels() -> Dict[str, str]:
    return {
        "cogs": "COGS (производство устройств)",
        "logistics": "Логистика",
        "support": "Поддержка пользователей",
        "infrastructure_variable": "Инфраструктура (переменная)",
        "cac": "CAC (привлечение клиентов)",
    }


# ──────────────────────────────────────────────
# Лист 1: Титульный лист
# ──────────────────────────────────────────────

def _build_title_sheet(wb: Workbook, model_type: str, all_params: Dict,
                        num_months: int, cash_flow_results: List[Dict],
                        breakeven_result: Dict, initial_investment: float) -> None:
    ws = wb.create_sheet("Титульный лист")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 20

    row = 1
    _merge(ws, row, 2, row, 4,
           "ФИНАНСОВАЯ МОДЕЛЬ ПРОЕКТА",
           font=Font(bold=True, size=16, color=_BLUE),
           align="center", fill=_HDR_FILL)
    ws.row_dimensions[row].height = 32
    row += 1

    _merge(ws, row, 2, row, 4,
           "ReFlex — аппаратно-программный комплекс для реабилитации ОДА",
           font=Font(bold=True, size=12, color=_WHITE),
           align="center",
           fill=PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid"))
    ws.row_dimensions[row].height = 22
    row += 2

    # Рекомендации из шаблона МСП
    notes = [
        "1. Это финансовая модель для конкурса МСП / Студенческий стартап, "
        "сгенерированная автоматически из калькулятора ReFlex ФЭМ.",
        "2. Данные приведены в тысячах рублей (тыс. руб.).",
        "3. Горизонт планирования: {} месяцев.".format(num_months),
        "4. Использованная бизнес-модель: {}".format(_model_name(model_type)),
    ]
    _merge(ws, row, 2, row, 4,
           "Рекомендации к финансовой модели:",
           font=_SECTION_FONT, fill=_SECTION_FILL, align="left")
    row += 1
    for note in notes:
        _merge(ws, row, 2, row, 4, note, font=_NORMAL_FONT, wrap=True)
        ws.row_dimensions[row].height = 18
        row += 1
    row += 1

    # Сводные показатели
    _merge(ws, row, 2, row, 4,
           "Сводные показатели проекта",
           font=_SECTION_FONT, fill=_SECTION_FILL)
    row += 1

    total_revenue = sum(_safe(r.get("revenue", 0)) for r in cash_flow_results)
    total_cf = sum(_safe(r.get("cash_flow", 0)) for r in cash_flow_results)
    be_month = breakeven_result.get("breakeven_month")
    be_str = f"Месяц {be_month}" if be_month else "Не достигается в горизонте"

    metrics = [
        ("Горизонт планирования", f"{num_months} месяцев"),
        ("Бизнес-модель", _model_name(model_type)),
        ("Суммарная выручка за период", f"{_rub(total_revenue):,.1f} тыс. ₽"),
        ("Суммарный CF за период", f"{_rub(total_cf):,.1f} тыс. ₽"),
        ("Точка безубыточности", be_str),
        ("Начальные инвестиции / грант", f"{_rub(initial_investment):,.1f} тыс. ₽"),
        ("Дата формирования", date.today().strftime("%d.%m.%Y")),
    ]
    for label, value in metrics:
        _w(ws, row, 2, label, bold=True, size=10, border=True)
        _w(ws, row, 3, value, size=10, border=True)
        row += 1

    row += 1
    _merge(ws, row, 2, row, 4,
           "Описание продукта",
           font=_SECTION_FONT, fill=_SECTION_FILL)
    row += 1
    desc = (
        "ReFlex — носимые сенсоры (ЭМГ + IMU) в связке с мобильным приложением для "
        "объективного контроля восстановления после травм ОДА. Система сравнивает движение "
        "пациента с эталоном в реальном времени и формирует отчёты для врача."
    )
    _merge(ws, row, 2, row + 2, 4, desc, font=_NORMAL_FONT, wrap=True,
           fill=_BLUE_FILL_P)
    ws.row_dimensions[row].height = 60
    row += 3

    row += 1
    _merge(ws, row, 2, row, 4,
           "Команда проекта",
           font=_SECTION_FONT, fill=_SECTION_FILL)
    row += 1
    team = [
        ("CEO / Руководитель проекта", "Антон"),
        ("MVP / Разработка", "Олег"),
        ("Дизайн / UX", "Э Рен"),
        ("Аналитика", "Мария"),
        ("GR / Юридика", "Павел"),
    ]
    for role, name in team:
        _w(ws, row, 2, role, bold=True, size=10, border=True)
        _w(ws, row, 3, name, size=10, border=True)
        row += 1


# ──────────────────────────────────────────────
# Лист 2: БДДС
# ──────────────────────────────────────────────

def _build_bdds_sheet(
    wb: Workbook,
    model_type: str,
    num_months: int,
    revenue_results: List[Dict],
    costs_results: List[Dict],
    cash_flow_results: List[Dict],
    bank_allocation: List[Dict],
    all_params: Dict,
    rnd_results: Optional[List[Dict]] = None,
) -> None:
    ws = wb.create_sheet("БДДС")
    ws.freeze_panes = "C3"

    rnd_count = len(rnd_results) if rnd_results else 0
    N = num_months + rnd_count  # total number of period columns
    # Column layout: A=label, B=sub-label, C..C+N-1=periods, C+N=ИТОГО
    total_col = 3 + N  # 1-indexed column for ИТОГО
    _first_period_col = 3       # column C
    _last_period_col = 2 + N    # last period column

    # ── Ширина столбцов ──
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 45
    for i in range(N + 1):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13

    # ── Строка 1: заголовок ──
    _w(ws, 1, 1, "БДР", bold=True, size=11)
    _w(ws, 1, 2, "тыс. руб.", italic=True, size=10)
    _w(ws, 1, total_col - N, "Видеоинструкция", italic=True, size=9, color="595959")

    # ── Строка 2: периоды (R&D месяцы + рыночные месяцы) ──
    _w(ws, 2, 2, "Период", font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)
    _rnd_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    _rnd_font = Font(bold=True, size=10, color="92400E")
    for r_idx in range(rnd_count):
        c = ws.cell(row=2, column=3 + r_idx, value=f"R&D {r_idx + 1}")
        c.fill = _rnd_fill
        c.font = _rnd_font
        c.alignment = Alignment(horizontal="center")
        c.border = _THIN_BORDER
    for m in range(1, num_months + 1):
        _w(ws, 2, 2 + rnd_count + m, m, font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)
    _w(ws, 2, total_col, "Итого", font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)

    # ── Вспомогательные функции ─────────────────────────────────────────────

    def _combined(rnd_vals: List[float], market_vals: List[float]) -> List[float]:
        """Склеивает R&D и рыночные значения в единый список длиной N."""
        return list(rnd_vals[:rnd_count]) + list(market_vals[:num_months])

    def _sum_formula(r: int) -> str:
        """Формула =SUM(C{r}:{last_period_col}{r}) для итогового столбца."""
        return f"=SUM({get_column_letter(_first_period_col)}{r}:{get_column_letter(_last_period_col)}{r})"

    def _row_data(values: List[float], label: str, sub_label: str = "",
                  row_num: int = 0, section: bool = False,
                  indent: int = 0, fill=None, bold: bool = False,
                  fmt: str = _NUM_FMT_DEC) -> int:
        """Пишет строку с данными. ИТОГО — формула =SUM(). Возвращает номер строки."""
        nonlocal _current_row
        r = _current_row
        lbl = ("  " * indent) + (sub_label or label)
        _w(ws, r, 2, lbl, bold=bold or section, fill=fill, border=True,
           size=10, wrap=False)
        for i, v in enumerate(values[:N]):
            cell_v = round(v, 2)
            cell = ws.cell(row=r, column=3 + i, value=cell_v if abs(cell_v) > 0.001 else None)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=bold, size=10)
            if fill:
                cell.fill = fill
            cell.border = _THIN_BORDER
        # ИТОГО = SUM-формула вместо вычисленного итога
        tot_cell = ws.cell(row=r, column=total_col, value=_sum_formula(r))
        tot_cell.number_format = fmt
        tot_cell.font = Font(bold=True, size=10)
        tot_cell.alignment = Alignment(horizontal="right")
        if fill:
            tot_cell.fill = fill
        tot_cell.border = _THIN_BORDER
        _current_row += 1
        return r

    def _formula_row(label: str, formula_fn,
                     total_formula: str = None,
                     sub_label: str = "", indent: int = 0,
                     fill=None, bold: bool = False,
                     fmt: str = _NUM_FMT_DEC) -> int:
        """
        Пишет строку, где каждая ячейка периода — Excel-формула formula_fn(col_letter, row).
        ИТОГО: total_formula или =SUM(). Возвращает номер строки.
        """
        nonlocal _current_row
        r = _current_row
        lbl = ("  " * indent) + (sub_label or label)
        _w(ws, r, 2, lbl, bold=bold, fill=fill, border=True, size=10, wrap=False)
        for i in range(N):
            col_letter = get_column_letter(3 + i)
            formula = formula_fn(col_letter)
            cell = ws.cell(row=r, column=3 + i, value=formula)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=bold, size=10)
            if fill:
                cell.fill = fill
            cell.border = _THIN_BORDER
        tot_val = total_formula if total_formula is not None else _sum_formula(r)
        tot_cell = ws.cell(row=r, column=total_col, value=tot_val)
        tot_cell.number_format = fmt
        tot_cell.font = Font(bold=True, size=10)
        tot_cell.alignment = Alignment(horizontal="right")
        if fill:
            tot_cell.fill = fill
        tot_cell.border = _THIN_BORDER
        _current_row += 1
        return r

    _current_row = 3

    # Вычисляем R&D расходы по месяцам (нули, если R&D не задан)
    _rnd_total_costs = [_rub(r["total_costs"]) for r in (rnd_results or [])]
    _rnd_zeros = [0.0] * rnd_count  # R&D не даёт выручки

    # ── СЕКЦИЯ: Денежный поток от операционной деятельности ──
    op_cf_market = [_safe(r.get("cash_flow", 0)) for r in cash_flow_results]
    op_cf_rnd = [-r["total_costs"] for r in (rnd_results or [])]
    op_cf_combined = _combined([_rub(v) for v in op_cf_rnd], [_rub(v) for v in op_cf_market])
    op_section_row = _row_data(op_cf_combined, "Денежный поток от операционной деятельности",
                               section=True, fill=_BLUE_FILL_P, bold=True)

    # Выручка от реализации (R&D: 0)
    revenue_monthly_market = [_rub(_safe(r.get("revenue", 0))) for r in cash_flow_results]
    revenue_monthly = _combined(_rnd_zeros, revenue_monthly_market)
    _w(ws, _current_row, 1, "1", bold=True, size=10)
    revenue_row = _row_data(revenue_monthly, "Выручка от реализации",
                            bold=True, fill=_GREEN_FILL_P)

    # Количество продаж (пациентов или device-месяцев), R&D: 0
    qty_monthly_market: List[float] = []
    for r in revenue_results:
        qty = r.get("num_patients", r.get("billable_patients", r.get("new_patients", 0)))
        qty_monthly_market.append(_safe(qty))
    qty_monthly = _combined(_rnd_zeros, qty_monthly_market)
    _row_data(qty_monthly, "Количество активных пациентов (ед.)",
              indent=1, fmt=_NUM_FMT)

    # Средняя цена за единицу, R&D: 0
    price_monthly_market: List[float] = []
    for i, r in enumerate(revenue_results):
        rev = _safe(r.get("total_revenue", 0))
        qty = _safe(r.get("num_patients", r.get("billable_patients", 1)))
        price_monthly_market.append(round(rev / max(qty, 1) / _K, 2))
    price_monthly = _combined(_rnd_zeros, price_monthly_market)
    _row_data(price_monthly, "Средняя цена / ед. (тыс. ₽)", indent=1, fmt=_NUM_FMT_DEC)

    # Операционные расходы (R&D: затраты R&D)
    opex_monthly_market = [
        _rub(_safe(cr.get("fixed_costs", {}).get("total", 0))
             + _safe(cr.get("variable_costs", {}).get("total", 0)))
        for cr in costs_results
    ]
    opex_monthly = _combined(_rnd_total_costs, opex_monthly_market)
    _w(ws, _current_row, 1, "2", bold=True, size=10)
    opex_row = _row_data(opex_monthly, "Операционные расходы",
                         bold=True, fill=_ORANGE_FILL_P)

    # Ретроактивно заменяем op_section_row на формулы = Выручка − Операционные расходы.
    # Строка op_section_row была записана раньше с Python-данными; теперь, зная
    # revenue_row и opex_row, переводим её ячейки в формульный вид.
    _op_itogo_col = get_column_letter(total_col)
    for _i in range(N):
        _cl = get_column_letter(3 + _i)
        _cell = ws.cell(row=op_section_row, column=3 + _i)
        _cell.value = f"={_cl}{revenue_row}-{_cl}{opex_row}"
    ws.cell(row=op_section_row, column=total_col).value = (
        f"={_op_itogo_col}{revenue_row}-{_op_itogo_col}{opex_row}"
    )

    # R&D расходы детализация (только если R&D включён)
    if rnd_results:
        _w(ws, _current_row, 1, "2.0", bold=True, size=9)
        _row_data(_rnd_total_costs + [0.0] * num_months, "R&D расходы (до старта продаж)",
                  section=True, fill=_YELLOW_FILL_P, bold=True, indent=1)
        all_rnd_cats: List[str] = []
        for r in rnd_results:
            for cat in r.get("breakdown", {}).keys():
                if cat not in all_rnd_cats:
                    all_rnd_cats.append(cat)
        for cat in all_rnd_cats:
            cat_vals = [_rub(r["breakdown"].get(cat, 0.0)) for r in rnd_results]
            cat_vals_combined = cat_vals + [0.0] * num_months
            if any(abs(v) > 0.001 for v in cat_vals_combined):
                _row_data(cat_vals_combined, cat, indent=2)

    # ── Переменные расходы (R&D: 0) ──
    var_monthly_market = [_rub(_safe(cr.get("variable_costs", {}).get("total", 0))) for cr in costs_results]
    var_monthly = _combined(_rnd_zeros, var_monthly_market)
    _w(ws, _current_row, 1, "2.1", bold=True, size=9)
    _row_data(var_monthly, "Переменные расходы (рыночная фаза)", section=True, fill=_YELLOW_FILL_P, bold=True, indent=1)

    var_labels = _variable_cost_labels()
    for key, label in var_labels.items():
        vals_market = [_rub(_safe(cr.get("variable_costs", {}).get(key, 0))) for cr in costs_results]
        vals = _combined(_rnd_zeros, vals_market)
        if any(abs(v) > 0.001 for v in vals):
            _row_data(vals, label, indent=2)

    custom_var = all_params.get("custom_variable_costs", {})
    for name, data in custom_var.items():
        if data.get("type") == "Единоразовая (месяц 1)":
            market_vals = [_rub(_safe(data.get("value", 0)))] + [0.0] * (num_months - 1)
        else:
            market_vals = [_rub(_safe(data.get("value", 0)))] * num_months
        vals = _combined(_rnd_zeros, market_vals)
        if any(abs(v) > 0.001 for v in vals):
            _row_data(vals, name, indent=2)

    # ── Постоянные расходы (R&D: 0) ──
    fix_monthly_market = [_rub(_safe(cr.get("fixed_costs", {}).get("total", 0))) for cr in costs_results]
    fix_monthly = _combined(_rnd_zeros, fix_monthly_market)
    _w(ws, _current_row, 1, "2.2", bold=True, size=9)
    _row_data(fix_monthly, "Постоянные расходы (рыночная фаза)", section=True, fill=_YELLOW_FILL_P, bold=True, indent=1)

    fix_labels = _fixed_cost_labels()
    for key, label in fix_labels.items():
        vals_market = [_rub(_safe(cr.get("fixed_costs", {}).get(key, 0))) for cr in costs_results]
        vals = _combined(_rnd_zeros, vals_market)
        if any(abs(v) > 0.001 for v in vals):
            _row_data(vals, label, indent=2)

    custom_fix = all_params.get("custom_fixed_costs", {})
    for name, data in custom_fix.items():
        if data.get("type") == "Единоразовая (месяц 1)":
            market_vals = [_rub(_safe(data.get("value", 0)))] + [0.0] * (num_months - 1)
        else:
            market_vals = [_rub(_safe(data.get("value", 0)))] * num_months
        vals = _combined(_rnd_zeros, market_vals)
        if any(abs(v) > 0.001 for v in vals):
            _row_data(vals, name, indent=2)

    # ── Прибыль до налогообложения = Выручка − Операционные расходы ──
    _w(ws, _current_row, 1, "3", bold=True, size=10)
    _itogo_col_letter = get_column_letter(total_col)
    profit_row = _formula_row(
        "Прибыль до налогообложения",
        formula_fn=lambda col: f"={col}{revenue_row}-{col}{opex_row}",
        total_formula=f"={_itogo_col_letter}{revenue_row}-{_itogo_col_letter}{opex_row}",
        bold=True, fill=_GREEN_FILL_P,
    )

    # Налог УСН 6% с выручки
    tax_row = _formula_row(
        "Налог УСН 6% (от выручки)",
        formula_fn=lambda col: f"={col}{revenue_row}*{_TAX_RATE}",
        total_formula=f"={_itogo_col_letter}{revenue_row}*{_TAX_RATE}",
        indent=1,
    )

    # Чистая прибыль = Прибыль − Налог
    _w(ws, _current_row, 1, "4", bold=True, size=10)
    net_row = _formula_row(
        "Чистая прибыль",
        formula_fn=lambda col: f"={col}{profit_row}-{col}{tax_row}",
        total_formula=f"={_itogo_col_letter}{profit_row}-{_itogo_col_letter}{tax_row}",
        bold=True, fill=_GREEN_FILL_P,
    )

    # Норма чистой прибыли = ЧП / Выручка
    _formula_row(
        "Норма чистой прибыли",
        formula_fn=lambda col: f"=IF({col}{revenue_row}<>0,{col}{net_row}/{col}{revenue_row},0)",
        total_formula=f"=IF({_itogo_col_letter}{revenue_row}<>0,{_itogo_col_letter}{net_row}/{_itogo_col_letter}{revenue_row},0)",
        indent=1, fmt=_PCT_FMT,
    )

    _current_row += 1  # разделитель

    # ── СЕКЦИЯ: Финансовая деятельность (грант) ──
    bank_rnd = [_rub(r["total_costs"]) for r in (rnd_results or [])]
    # bank_allocation для рыночных месяцев: первые rnd_count записей — R&D (если combined_bank_allocation)
    # Берём только маркетные записи из bank_allocation (пропускаем R&D-записи)
    _market_bank_alloc = [e for e in (bank_allocation or []) if not str(e.get("month", "")).startswith("R&D")]
    bank_market = [_rub(_sum_bank_by_month(_market_bank_alloc, i)) for i in range(num_months)]
    bank_monthly = _combined(bank_rnd, bank_market)
    _w(ws, _current_row, 1, "5", bold=True, size=10)
    bank_section_row = _row_data(bank_monthly, "Денежный поток от финансовой деятельности",
                                 section=True, fill=_BLUE_FILL_P, bold=True)

    _row_data(bank_monthly, "Расходование грантовых / инвестиционных средств", indent=1)

    if rnd_results:
        all_rnd_cats_bank: List[str] = []
        for r in rnd_results:
            for cat in r.get("breakdown", {}).keys():
                if cat not in all_rnd_cats_bank:
                    all_rnd_cats_bank.append(cat)
        for cat in all_rnd_cats_bank:
            cat_rnd = [_rub(r["breakdown"].get(cat, 0.0)) for r in rnd_results]
            cat_vals = cat_rnd + [0.0] * num_months
            if any(abs(v) > 0.001 for v in cat_vals):
                _row_data(cat_vals, f"{cat} (R&D, из гранта)", indent=2)

    cat_labels = {
        "team_salaries": "Зарплаты (из гранта)",
        "infrastructure_fixed": "Инфраструктура (из гранта)",
        "office_rent": "Аренда офиса (из гранта)",
        "legal_services": "Юридические услуги (из гранта)",
        "other_fixed": "Прочие постоянные (из гранта)",
        "cogs": "COGS (из гранта)",
        "logistics": "Логистика (из гранта)",
        "support": "Поддержка (из гранта)",
        "infrastructure_variable": "Инфраструктура перем. (из гранта)",
        "cac": "CAC (из гранта)",
    }
    for cat, lbl in cat_labels.items():
        cat_market = []
        for i in range(num_months):
            entry = _market_bank_alloc[i] if i < len(_market_bank_alloc) else {}
            line_items = entry.get("line_items", {})
            v = line_items.get(cat, {}).get("bank", 0.0) if isinstance(line_items.get(cat), dict) else _safe(entry.get(cat, 0))
            cat_market.append(_rub(v))
        cat_vals = _combined(_rnd_zeros, cat_market)
        if any(abs(v) > 0.001 for v in cat_vals):
            _row_data(cat_vals, lbl, indent=2)

    _current_row += 1  # разделитель

    # ── СЕКЦИЯ: Остатки денежных средств ──
    _w(ws, _current_row, 1, "6", bold=True, size=10)

    # Оборот = op_cf + bank (для R&D: op_cf=-rnd, bank=+rnd → net≈0)
    # Вычисляем opening через cumulative для корректной инициализации
    rnd_cum_cf = [_rub(r["cumulative_cash_flow"]) for r in (rnd_results or [])]
    market_cum_cf = [_rub(_safe(r.get("cumulative_cash_flow", 0))) for r in cash_flow_results]
    rnd_cum_last = rnd_cum_cf[-1] if rnd_cum_cf else 0.0
    market_cum_cf_adj = [rnd_cum_last + v for v in market_cum_cf]
    cum_cf = rnd_cum_cf + market_cum_cf_adj
    opening_vals = [0.0] + [cum_cf[i] for i in range(N - 1)]
    opening_row = _row_data(opening_vals, "Остаток ДС на начало периода",
                            section=True, fill=_GRAY_FILL_P, bold=True)

    # Оборот = Операционный CF + Финансовый CF (формула ссылается на строки op_section и bank_section)
    turnover_row = _formula_row(
        "Оборот за период (CF + финансирование)",
        formula_fn=lambda col: f"={col}{op_section_row}+{col}{bank_section_row}",
        indent=1,
    )

    # Остаток на конец = Остаток на начало + Оборот
    _formula_row(
        "Остаток ДС на конец периода",
        formula_fn=lambda col: f"={col}{opening_row}+{col}{turnover_row}",
        total_formula=_sum_formula(_current_row),  # placeholder, будет перезаписано
        bold=True, fill=_GREEN_FILL_P,
    )
    # Исправляем total для closing: итог = последний элемент (не сумма)
    _closing_row = _current_row - 1
    _last_col_letter = get_column_letter(_last_period_col)
    ws.cell(row=_closing_row, column=total_col,
            value=f"={_last_col_letter}{_closing_row}").number_format = _NUM_FMT_DEC

    # Ретроактивно обновляем opening_row: периоды 2..N ссылаются на closing предыдущего периода.
    # Период 1 (col C) остаётся = 0 (начальный остаток).
    for _i in range(1, N):
        _prev_col = get_column_letter(3 + _i - 1)
        _cell = ws.cell(row=opening_row, column=3 + _i)
        _cell.value = f"={_prev_col}{_closing_row}"

    # ── Заморозка и финальная ширина ──
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 20


# ──────────────────────────────────────────────
# Лист 3: Инвест
# ──────────────────────────────────────────────

def _build_invest_sheet(
    wb: Workbook,
    num_months: int,
    cash_flow_results: List[Dict],
    revenue_results: List[Dict],
    costs_results: List[Dict],
    bank_allocation: List[Dict],
    breakeven_result: Dict,
    initial_investment: float,
    discount_rate_annual: float,
    rnd_results: Optional[List[Dict]] = None,
) -> None:
    ws = wb.create_sheet("Инвест")
    # Строки 3–4: параметры дисконтирования (как в калькуляторе: годовая → месячная r_m).
    # Таблица показателей начинается со строки 5.
    ws.freeze_panes = "C5"

    N = num_months
    total_col = 3 + N
    _first_col = 3
    _last_col = 2 + N

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 45
    for i in range(N + 1):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14

    # ── Заголовок ──
    _merge(ws, 1, 2, 1, total_col,
           "ИНВЕСТИЦИОННЫЙ АНАЛИЗ — тыс. руб.",
           font=_TITLE_FONT, fill=_HDR_FILL, align="center")
    _w(ws, 1, 1, "", fill=_HDR_FILL)

    _w(ws, 2, 2, "", font=_HDR_FONT, fill=_HDR_FILL)
    for m in range(1, N + 1):
        _w(ws, 2, 2 + m, m, font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)
    _w(ws, 2, total_col, "Итого / Среднее", font=_HDR_FONT, fill=_HDR_FILL,
        align="center", border=True)

    # Параметр из калькулятора (левая панель): годовая ставка дисконтирования.
    # Месячная ставка в C4 совпадает с models/cash_flow.py: r_m = (1+r_год)^(1/12)-1
    _r_annual = float(discount_rate_annual) if discount_rate_annual is not None else 0.0
    if _r_annual <= -1.0:
        _r_annual = 0.0
    _w(ws, 3, 2,
        "Годовая ставка дисконтирования (из калькулятора, доля: 0,20 = 20%)",
        size=10, border=True, fill=_BLUE_FILL_P)
    _c3 = ws.cell(row=3, column=3, value=_r_annual)
    _c3.number_format = "0.00%"
    _c3.font = Font(bold=True, size=10)
    _c3.alignment = Alignment(horizontal="right")
    _c3.border = _THIN_BORDER
    _w(ws, 4, 2,
        "Месячная ставка r_m = (1+r_год)^(1/12)-1 (как в NPV калькулятора)",
        size=10, border=True, fill=_GRAY_FILL_P)
    _c4 = ws.cell(row=4, column=3, value="=(1+C3)^(1/12)-1")
    _c4.number_format = _NUM_FMT_DEC
    _c4.font = Font(bold=True, size=10, color=_BLUE)
    _c4.alignment = Alignment(horizontal="right")
    _c4.border = _THIN_BORDER
    _r_m_cell_abs = "$C$4"  # абсолютная ссылка для формул DCF

    def _sum_inv(r: int) -> str:
        return f"=SUM({get_column_letter(_first_col)}{r}:{get_column_letter(_last_col)}{r})"

    def _inv_row(row: int, label: str, values: List[float],
                 bold: bool = False, fill=None, fmt=_NUM_FMT_DEC,
                 total_formula: str = None) -> int:
        """Пишет строку с данными. ИТОГО — формула =SUM(). Возвращает row."""
        _w(ws, row, 2, label, bold=bold, size=10, fill=fill, border=True)
        for i, v in enumerate(values[:N]):
            cell = ws.cell(row=row, column=3 + i, value=round(v, 2) if abs(v) > 0.001 else None)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=bold, size=10)
            if fill:
                cell.fill = fill
            cell.border = _THIN_BORDER
        tot_val = total_formula if total_formula is not None else _sum_inv(row)
        tot = ws.cell(row=row, column=total_col, value=tot_val)
        tot.number_format = fmt
        tot.font = Font(bold=True, size=10)
        tot.alignment = Alignment(horizontal="right")
        if fill:
            tot.fill = fill
        tot.border = _THIN_BORDER
        return row

    def _formula_inv(row: int, label: str, formula_fn,
                     bold: bool = False, fill=None, fmt=_NUM_FMT_DEC,
                     total_formula: str = None) -> int:
        """Пишет строку с Excel-формулами. Возвращает row."""
        _w(ws, row, 2, label, bold=bold, size=10, fill=fill, border=True)
        for i in range(N):
            col_letter = get_column_letter(3 + i)
            cell = ws.cell(row=row, column=3 + i, value=formula_fn(col_letter))
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
            cell.font = Font(bold=bold, size=10)
            if fill:
                cell.fill = fill
            cell.border = _THIN_BORDER
        tot_val = total_formula if total_formula is not None else _sum_inv(row)
        tot = ws.cell(row=row, column=total_col, value=tot_val)
        tot.number_format = fmt
        tot.font = Font(bold=True, size=10)
        tot.alignment = Alignment(horizontal="right")
        if fill:
            tot.fill = fill
        tot.border = _THIN_BORDER
        return row

    row = 5
    _tc = get_column_letter(total_col)  # letter of ИТОГО column

    # ── 1. Выручка ──
    rev_m = [_rub(_safe(r.get("revenue", 0))) for r in cash_flow_results]
    _w(ws, row, 1, "1", bold=True)
    rev_row = _inv_row(row, "Выручка", rev_m, bold=True, fill=_GREEN_FILL_P)
    row += 1

    # ── 2. Переменные расходы ──
    var_m = [_rub(_safe(cr.get("variable_costs", {}).get("total", 0))) for cr in costs_results]
    _w(ws, row, 1, "2", bold=True)
    var_row = _inv_row(row, "Переменные расходы", var_m, fill=_ORANGE_FILL_P)
    row += 1

    # ── 3. Постоянные расходы ──
    fix_m = [_rub(_safe(cr.get("fixed_costs", {}).get("total", 0))) for cr in costs_results]
    _w(ws, row, 1, "3", bold=True)
    fix_row = _inv_row(row, "Постоянные расходы", fix_m, fill=_ORANGE_FILL_P)
    row += 1

    # ── 4. Прибыль до НО = Выручка − Перем.расх − Пост.расх ──
    _w(ws, row, 1, "4", bold=True)
    profit_row = _formula_inv(
        row, "Прибыль до налогообложения",
        formula_fn=lambda col: f"={col}{rev_row}-{col}{var_row}-{col}{fix_row}",
        total_formula=f"={_tc}{rev_row}-{_tc}{var_row}-{_tc}{fix_row}",
        bold=True, fill=_YELLOW_FILL_P,
    )
    row += 1

    # ── 5. Чистая прибыль = Прибыль − Выручка × 6% ──
    _w(ws, row, 1, "5", bold=True)
    net_row = _formula_inv(
        row, "Чистая прибыль",
        formula_fn=lambda col: f"={col}{profit_row}-{col}{rev_row}*{_TAX_RATE}",
        total_formula=f"={_tc}{profit_row}-{_tc}{rev_row}*{_TAX_RATE}",
        bold=True, fill=_GREEN_FILL_P,
    )
    row += 1

    # ── Норма ЧП = ЧП / Выручка ──
    _formula_inv(
        row, "Норма чистой прибыли",
        formula_fn=lambda col: f"=IF({col}{rev_row}<>0,{col}{net_row}/{col}{rev_row},0)",
        total_formula=f"=IF({_tc}{rev_row}<>0,{_tc}{net_row}/{_tc}{rev_row},0)",
        fmt=_PCT_FMT,
    )
    row += 1

    # ── 6. Инвестиции / грант (только рыночные месяцы) ──
    _market_bank_alloc_inv = [e for e in (bank_allocation or []) if not str(e.get("month", "")).startswith("R&D")]
    inv_m = [_rub(_sum_bank_by_month(_market_bank_alloc_inv, i)) for i in range(N)]
    _w(ws, row, 1, "6", bold=True)
    _inv_row(row, "Инвестиции (расходование гранта)", inv_m, fill=_BLUE_FILL_P)
    row += 1

    # ── 7. Денежный поток (CF) = Выручка − Перем.расх − Пост.расх ──
    _w(ws, row, 1, "7", bold=True)
    cf_row = _formula_inv(
        row, "Денежный поток (CF)",
        formula_fn=lambda col: f"={col}{rev_row}-{col}{var_row}-{col}{fix_row}",
        total_formula=f"={_tc}{rev_row}-{_tc}{var_row}-{_tc}{fix_row}",
        bold=True, fill=_YELLOW_FILL_P,
    )
    row += 1

    # ── 8. DCF = CF / (1 + r_m)^(COLUMN()-2+rnd_offset) ──
    # r_m берётся из ячейки C4 = (1+C3)^(1/12)-1 (совпадает с calculate_npv_series).
    # COLUMN()-2 — порядковый номер рыночного месяца (1 для C, 2 для D, ...).
    # rnd_offset сдвигает экспоненту на длину R&D фазы (как month_offset в калькуляторе).
    _rnd_count_invest = len(rnd_results) if rnd_results else 0
    _w(ws, row, 1, "8", bold=True)
    dcf_label = (
        "DCF (дисконтированный CF по r_m из C4"
        + (f", R&D offset={_rnd_count_invest} мес" if _rnd_count_invest else "")
        + ")"
    )
    dcf_row_idx = _formula_inv(
        row, dcf_label,
        formula_fn=lambda col: (
            f"={col}{cf_row}/(1+{_r_m_cell_abs})^(COLUMN()-2+{_rnd_count_invest})"
        ),
        fill=_YELLOW_FILL_P,
    )
    row += 1

    # ── ADCF — накопленный дисконтированный CF: =SUM($C{dcf}:{col}{dcf}) ──
    adcf_row_idx = row
    _w(ws, row, 2, "ADCF (накопленный DCF)", size=10, fill=_YELLOW_FILL_P, border=True)
    for i in range(N):
        col_letter = get_column_letter(3 + i)
        # Абсолютный старт ($C), относительный конец (текущий столбец)
        formula = f"=SUM($C{dcf_row_idx}:{col_letter}{dcf_row_idx})"
        cell = ws.cell(row=row, column=3 + i, value=formula)
        cell.number_format = _NUM_FMT_DEC
        cell.alignment = Alignment(horizontal="right")
        cell.font = Font(size=10)
        cell.fill = _YELLOW_FILL_P
        cell.border = _THIN_BORDER
    # ИТОГО для ADCF = последний элемент (накопленный итог)
    _last_adcf_col = get_column_letter(_last_col)
    tot_adcf = ws.cell(row=row, column=total_col,
                       value=f"={_last_adcf_col}{row}")
    tot_adcf.number_format = _NUM_FMT_DEC
    tot_adcf.font = Font(bold=True, size=10)
    tot_adcf.alignment = Alignment(horizontal="right")
    tot_adcf.fill = _YELLOW_FILL_P
    tot_adcf.border = _THIN_BORDER
    row += 2

    # ── Параметры CAPM ──
    _merge(ws, row, 2, row, total_col,
           "Параметры ставки дисконтирования (CAPM)",
           font=_SECTION_FONT, fill=_SECTION_FILL)
    row += 1

    capm_params = [
        ("r (ставка дисконтирования)", f"{_CAPM_RATE * 100:.2f}%"),
        ("Безрисковая ставка (r_f)", f"{_RISK_FREE * 100:.1f}%"),
        ("Бета (β)", f"{_BETA}"),
        ("Среднерыночный доход (r_m)", f"{_MARKET_RETURN * 100:.2f}%"),
    ]
    for label, val in capm_params:
        _w(ws, row, 2, label, size=10, border=True)
        _w(ws, row, 3, val, size=10, border=True)
        row += 1

    # ── Строка с I₀ (начальные инвестиции) — даём ей именованный диапазон ──
    i0_row = row + 1  # будет записан ниже в метриках
    row += 1

    # ── Итоговые инвестиционные метрики ──
    _merge(ws, row, 2, row, total_col,
           "Ключевые инвестиционные метрики",
           font=_SECTION_FONT, fill=_SECTION_FILL)
    row += 1

    be_month = breakeven_result.get("breakeven_month")

    # NPV = ИТОГО DCF (формула ссылается на total_col строки DCF)
    npv_row = row
    _w(ws, row, 2, "NPV (сумма DCF), тыс. ₽", bold=True, size=10, border=True, fill=_BLUE_FILL_P)
    npv_cell = ws.cell(row=row, column=3,
                       value=f"={get_column_letter(total_col)}{dcf_row_idx}")
    npv_cell.number_format = _NUM_FMT_DEC
    npv_cell.font = Font(bold=True, size=11, color=_BLUE)
    npv_cell.alignment = Alignment(horizontal="center")
    npv_cell.border = _THIN_BORDER
    row += 1

    # NPV накопленным итогом = ИТОГО ADCF (формула ссылается на total_col ADCF)
    _w(ws, row, 2, "NPV нарастающим итогом, тыс. ₽", bold=True, size=10, border=True, fill=_BLUE_FILL_P)
    npv_cum_cell = ws.cell(row=row, column=3,
                           value=f"={get_column_letter(total_col)}{adcf_row_idx}")
    npv_cum_cell.number_format = _NUM_FMT_DEC
    npv_cum_cell.font = Font(bold=True, size=11, color=_BLUE)
    npv_cum_cell.alignment = Alignment(horizontal="center")
    npv_cum_cell.border = _THIN_BORDER
    row += 1

    # PP
    _w(ws, row, 2, "PP — срок окупаемости (месяц)", bold=True, size=10, border=True, fill=_BLUE_FILL_P)
    pp_cell = ws.cell(row=row, column=3, value=be_month if be_month else "Не достигается")
    pp_cell.font = Font(bold=True, size=11, color=_BLUE)
    pp_cell.alignment = Alignment(horizontal="center")
    pp_cell.border = _THIN_BORDER
    row += 1

    # I₀ — начальные инвестиции (вводимый параметр, не формула)
    i0_row = row
    _w(ws, row, 2, "Начальные инвестиции, тыс. ₽", bold=True, size=10, border=True, fill=_BLUE_FILL_P)
    i0_cell = ws.cell(row=row, column=3, value=round(initial_investment / _K, 2))
    i0_cell.number_format = _NUM_FMT_DEC
    i0_cell.font = Font(bold=True, size=11, color=_BLUE)
    i0_cell.alignment = Alignment(horizontal="center")
    i0_cell.border = _THIN_BORDER
    row += 1

    # PI = NPV / I₀ (формула ссылается на ячейки выше)
    _w(ws, row, 2, "PI — индекс прибыльности", bold=True, size=10, border=True, fill=_BLUE_FILL_P)
    pi_cell = ws.cell(row=row, column=3,
                      value=f"=IF(C{i0_row}<>0,C{npv_row}/C{i0_row},0)")
    pi_cell.number_format = _NUM_FMT_DEC
    pi_cell.font = Font(bold=True, size=11, color=_BLUE)
    pi_cell.alignment = Alignment(horizontal="center")
    pi_cell.border = _THIN_BORDER
    row += 1


# ──────────────────────────────────────────────
# Лист 4: График денежных потоков (График 1)
# ──────────────────────────────────────────────

def _build_cf_chart_sheet(
    wb: Workbook,
    num_months: int,
    cash_flow_results: List[Dict],
    bank_allocation: List[Dict],
) -> None:
    ws = wb.create_sheet("График ДП")

    N = num_months
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    _merge(ws, 1, 1, 1, 5,
           "ГРАФИК 1: ДЕНЕЖНЫЕ ПОТОКИ ПО МЕСЯЦАМ",
           font=_TITLE_FONT, fill=_HDR_FILL, align="center")

    # Заголовки таблицы
    headers = ["Месяц", "Выручка (тыс.₽)", "Расходы (тыс.₽)", "CF мес. (тыс.₽)", "CF накопл. (тыс.₽)"]
    for i, h in enumerate(headers, start=1):
        _w(ws, 2, i, h, font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)

    # Данные
    cum = 0.0
    for m in range(N):
        row = 3 + m
        cf_row = cash_flow_results[m]
        rev = _rub(_safe(cf_row.get("revenue", 0)))
        costs = _rub(_safe(cf_row.get("total_costs", 0)))
        cf_val = _rub(_safe(cf_row.get("cash_flow", 0)))
        cum += cf_val
        is_pos = cf_val >= 0

        _w(ws, row, 1, m + 1, align="center", border=True)
        _w(ws, row, 2, rev, fmt=_NUM_FMT_DEC, align="right", border=True)
        _w(ws, row, 3, costs, fmt=_NUM_FMT_DEC, align="right", border=True)
        fill = _GREEN_FILL_P if is_pos else _ORANGE_FILL_P
        _w(ws, row, 4, cf_val, fmt=_NUM_FMT_DEC, align="right", border=True, fill=fill)
        cum_fill = _GREEN_FILL_P if cum >= 0 else _ORANGE_FILL_P
        _w(ws, row, 5, round(cum, 2), fmt=_NUM_FMT_DEC, align="right", border=True, fill=cum_fill)

    # ── BarChart: CF по месяцам ──
    bar = BarChart()
    bar.type = "col"
    bar.grouping = "clustered"
    bar.title = "Денежный поток по месяцам"
    bar.y_axis.title = "тыс. ₽"
    bar.x_axis.title = "Месяц"
    bar.style = 10
    bar.width = 20
    bar.height = 12

    # Выручка
    rev_ref = Reference(ws, min_col=2, min_row=2, max_row=2 + N)
    bar.add_data(rev_ref, titles_from_data=True)

    # Расходы
    cost_ref = Reference(ws, min_col=3, min_row=2, max_row=2 + N)
    bar.add_data(cost_ref, titles_from_data=True)

    # CF
    cf_ref = Reference(ws, min_col=4, min_row=2, max_row=2 + N)
    bar.add_data(cf_ref, titles_from_data=True)

    cats = Reference(ws, min_col=1, min_row=3, max_row=2 + N)
    bar.set_categories(cats)
    ws.add_chart(bar, f"G2")

    # ── LineChart: накопленный CF ──
    line = LineChart()
    line.title = "Накопленный денежный поток"
    line.y_axis.title = "тыс. ₽ (накопл.)"
    line.x_axis.title = "Месяц"
    line.style = 10
    line.width = 20
    line.height = 12

    cum_ref = Reference(ws, min_col=5, min_row=2, max_row=2 + N)
    line.add_data(cum_ref, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, f"G{2 + 22}")


# ──────────────────────────────────────────────
# Лист 5: Точка безубыточности
# ──────────────────────────────────────────────

def _build_breakeven_sheet(
    wb: Workbook,
    num_months: int,
    cash_flow_results: List[Dict],
    costs_results: List[Dict],
    breakeven_result: Dict,
) -> None:
    ws = wb.create_sheet("Точка безубыточности")

    N = num_months
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18

    _merge(ws, 1, 1, 1, 4,
           "АНАЛИЗ ТОЧКИ БЕЗУБЫТОЧНОСТИ",
           font=_TITLE_FONT, fill=_HDR_FILL, align="center")

    # Резюме
    be_month = breakeven_result.get("breakeven_month")
    be_str = f"Месяц {be_month}" if be_month else "В горизонте не достигается"
    row = 2
    _merge(ws, row, 1, row, 4,
           f"Точка безубыточности: {be_str}",
           font=Font(bold=True, size=12, color=_BLUE if be_month else _RED),
           fill=_GREEN_FILL_P if be_month else _ORANGE_FILL_P, align="center")
    row += 1

    headers = ["Месяц", "Выручка (тыс.₽)", "Перем. затраты (тыс.₽)",
               "Пост. затраты (тыс.₽)", "Всего затраты (тыс.₽)", "Маржин. прибыль (тыс.₽)"]
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 20
    for i, h in enumerate(headers, start=1):
        _w(ws, row, i, h, font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)
    row += 1
    data_start = row

    for m in range(N):
        cr = costs_results[m]
        cfr = cash_flow_results[m]
        rev = _rub(_safe(cfr.get("revenue", 0)))
        var_c = _rub(_safe(cr.get("variable_costs", {}).get("total", 0)))
        fix_c = _rub(_safe(cr.get("fixed_costs", {}).get("total", 0)))
        total_c = var_c + fix_c
        margin = rev - var_c

        fill = _GREEN_FILL_P if rev >= total_c else _ORANGE_FILL_P

        _w(ws, row, 1, m + 1, align="center", border=True)
        _w(ws, row, 2, rev, fmt=_NUM_FMT_DEC, align="right", border=True)
        _w(ws, row, 3, var_c, fmt=_NUM_FMT_DEC, align="right", border=True)
        _w(ws, row, 4, fix_c, fmt=_NUM_FMT_DEC, align="right", border=True)
        _w(ws, row, 5, total_c, fmt=_NUM_FMT_DEC, align="right", border=True, fill=fill)
        _w(ws, row, 6, margin, fmt=_NUM_FMT_DEC, align="right", border=True)
        row += 1

    data_end = row - 1

    # ── LineChart: Revenue vs Total Costs vs Fixed Costs ──
    chart = LineChart()
    chart.title = "Точка безубыточности: Выручка vs Затраты"
    chart.y_axis.title = "тыс. ₽"
    chart.x_axis.title = "Месяц"
    chart.style = 10
    chart.width = 24
    chart.height = 14

    rev_ref = Reference(ws, min_col=2, min_row=data_start - 1, max_row=data_end)
    chart.add_data(rev_ref, titles_from_data=True)

    total_c_ref = Reference(ws, min_col=5, min_row=data_start - 1, max_row=data_end)
    chart.add_data(total_c_ref, titles_from_data=True)

    fix_c_ref = Reference(ws, min_col=4, min_row=data_start - 1, max_row=data_end)
    chart.add_data(fix_c_ref, titles_from_data=True)

    cats = Reference(ws, min_col=1, min_row=data_start, max_row=data_end)
    chart.set_categories(cats)
    ws.add_chart(chart, "H3")


# ──────────────────────────────────────────────
# Лист 6: Когорты пациентов
# ──────────────────────────────────────────────

def _build_cohorts_sheet(
    wb: Workbook,
    num_months: int,
    revenue_results: List[Dict],
    all_params: Dict,
) -> None:
    ws = wb.create_sheet("Когорты пациентов")

    N = num_months
    rev_params = all_params.get("revenue", {})
    rehab_dur = int(rev_params.get("rehab_duration_months", 3))

    ws.column_dimensions["A"].width = 22
    for i in range(N + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 10

    _merge(ws, 1, 1, 1, min(N + 2, 26),
           "КОГОРТНЫЙ АНАЛИЗ ПАЦИЕНТОВ",
           font=_TITLE_FONT, fill=_HDR_FILL, align="center")

    row = 2
    _merge(ws, row, 1, row, min(N + 2, 26),
           f"Длительность реабилитационного курса: {rehab_dur} мес. | "
           f"Каждая строка = новая когорта пациентов, стартовавшая в данном месяце.",
           font=_ITALIC_FONT, fill=_GRAY_FILL_P, wrap=True)
    ws.row_dimensions[row].height = 28
    row += 1

    # Заголовки: Когорта \ Месяц активности
    _w(ws, row, 1, "Когорта (старт)", font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)
    for m in range(1, N + 1):
        _w(ws, row, 1 + m, f"М{m}", font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)
    _w(ws, row, N + 2, "Всего за курс", font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)
    row += 1

    data_start = row
    # Строим когортную матрицу на основе new_patients и rehab_duration
    new_patients = [int(_safe(r.get("new_patients", 0))) for r in revenue_results]

    for cohort_month in range(N):
        cohort_size = new_patients[cohort_month]
        label = f"Когорта М{cohort_month + 1} ({cohort_size} пац.)"
        _w(ws, row, 1, label, bold=True, border=True, size=9)
        cohort_total = 0
        for act_month in range(N):
            active_in = (
                cohort_size
                if cohort_month <= act_month < cohort_month + rehab_dur
                else 0
            )
            fill = _GREEN_FILL_P if active_in > 0 else None
            cell = ws.cell(row=row, column=2 + act_month, value=active_in if active_in else None)
            cell.number_format = _NUM_FMT
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(size=9)
            cell.border = _THIN_BORDER
            if fill:
                cell.fill = fill
            cohort_total += active_in
        _w(ws, row, N + 2, cohort_total, fmt=_NUM_FMT, bold=True, align="right", border=True)
        row += 1

    # Итоговая строка: активные пациенты по месяцам (сумма когорт)
    _w(ws, row, 1, "ИТОГО активных пациентов", bold=True, fill=_BLUE_FILL_P, border=True)
    grand_total = 0
    for m in range(N):
        actual = int(_safe(revenue_results[m].get("num_patients", 0)))
        cell = ws.cell(row=row, column=2 + m, value=actual if actual else None)
        cell.number_format = _NUM_FMT
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = _BLUE_FILL_P
        cell.border = _THIN_BORDER
        grand_total += actual
    _w(ws, row, N + 2, grand_total, bold=True, fmt=_NUM_FMT, fill=_BLUE_FILL_P, border=True)
    data_end = row
    row += 2

    # ── Stacked Bar Chart ──
    if N <= 24:  # Ограничение видимости
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "stacked"
        chart.title = "Когортная динамика пациентов"
        chart.y_axis.title = "Кол-во пациентов"
        chart.x_axis.title = "Месяц"
        chart.style = 10
        chart.width = 26
        chart.height = 14
        chart.overlap = 100

        # Добавляем данные: итоговую строку активных пациентов
        act_ref = Reference(ws, min_col=2, min_row=data_end, max_col=1 + N)
        chart.add_data(act_ref)
        cats = Reference(ws, min_col=2, min_row=3, max_col=1 + N)
        chart.set_categories(cats)
        ws.add_chart(chart, f"A{row}")


# ──────────────────────────────────────────────
# Лист 7: Клиники
# ──────────────────────────────────────────────

def _build_clinics_sheet(
    wb: Workbook,
    num_months: int,
    revenue_results: List[Dict],
    all_params: Dict,
    clinic_schedule: List[Dict],
    model_type: str,
) -> None:
    ws = wb.create_sheet("Клиники")

    N = num_months
    rev_params = all_params.get("revenue", {})

    ws.column_dimensions["A"].width = 30
    for i in range(N + 1):
        ws.column_dimensions[get_column_letter(2 + i)].width = 12

    _merge(ws, 1, 1, 1, min(N + 2, 26),
           "ТАБЛИЦА: ПОЛЬЗОВАТЕЛИ В КЛИНИКАХ И ПОДКЛЮЧЕНИЕ КЛИНИК",
           font=_TITLE_FONT, fill=_HDR_FILL, align="center")

    # ── Секция 1: Подключение клиник ──
    row = 2
    _merge(ws, row, 1, row, min(N + 2, 26),
           "Секция 1: Расписание подключения клиник",
           font=_SECTION_FONT, fill=_SECTION_FILL)
    row += 1

    # Строим таблицу подключений
    # initial clinics + clinic_schedule
    initial_clinics = int(rev_params.get("num_clinics", 1))
    schedule_entries = [{"month_start": 1, "count": initial_clinics}]
    for entry in (clinic_schedule or []):
        ms = int(entry.get("month_start", 2))
        cnt = int(entry.get("count", 0))
        if ms >= 1 and cnt > 0:
            schedule_entries.append({"month_start": ms, "count": cnt})
    schedule_entries.sort(key=lambda x: x["month_start"])

    sched_headers = ["Месяц подключения", "Кол-во новых клиник", "Кумулятивно клиник", "Статус"]
    for i, h in enumerate(sched_headers, start=1):
        _w(ws, row, i, h, font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)
    row += 1

    cum_clinics = 0
    for entry in schedule_entries:
        ms = entry["month_start"]
        cnt = entry["count"]
        cum_clinics += cnt
        status = "Пилотный запуск" if ms == 1 else "Масштабирование"
        fill = _GREEN_FILL_P if ms == 1 else _BLUE_FILL_P
        _w(ws, row, 1, f"Месяц {ms}", align="center", border=True, fill=fill)
        _w(ws, row, 2, cnt, align="center", border=True, bold=True, fill=fill)
        _w(ws, row, 3, cum_clinics, align="center", border=True, bold=True, fill=fill)
        _w(ws, row, 4, status, align="center", border=True, fill=fill)
        row += 1
    row += 1

    # ── Секция 2: Пациенты в клиниках по месяцам ──
    _merge(ws, row, 1, row, min(N + 2, 26),
           "Секция 2: Пользователи в клиниках (активные пациенты по месяцам)",
           font=_SECTION_FONT, fill=_SECTION_FILL)
    row += 1

    _w(ws, row, 1, "Показатель", font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)
    for m in range(1, N + 1):
        _w(ws, row, 1 + m, f"М{m}", font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)
    _w(ws, row, N + 2, "Сумм.", font=_HDR_FONT, fill=_HDR_FILL, align="center", border=True)
    row += 1

    def _clinic_row(label: str, values: List[float], bold: bool = False,
                    fill=None, fmt=_NUM_FMT):
        _w(ws, row, 1, label, bold=bold, border=True, fill=fill)
        total = 0.0
        for i, v in enumerate(values[:N]):
            cell = ws.cell(row=row, column=2 + i, value=round(v, 1) if abs(v) > 0.01 else None)
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="center")
            cell.font = Font(bold=bold, size=10)
            cell.border = _THIN_BORDER
            if fill:
                cell.fill = fill
            total += v
        tot = ws.cell(row=row, column=N + 2, value=round(total, 1))
        tot.number_format = fmt
        tot.font = Font(bold=True, size=10)
        tot.alignment = Alignment(horizontal="center")
        tot.border = _THIN_BORDER
        if fill:
            tot.fill = fill

    # Число клиник по месяцам (нарастающим итогом)
    clinics_by_month: List[float] = []
    cum_c = 0
    sched_map: Dict[int, int] = {}
    for e in schedule_entries:
        sched_map[e["month_start"]] = sched_map.get(e["month_start"], 0) + e["count"]
    for m in range(1, N + 1):
        cum_c += sched_map.get(m, 0)
        clinics_by_month.append(float(cum_c))
    _clinic_row("Число подключённых клиник (накопл.)", clinics_by_month, bold=True, fill=_BLUE_FILL_P)
    row += 1

    # Активные пациенты всего
    active_patients = [float(_safe(r.get("num_patients", 0))) for r in revenue_results]
    _clinic_row("Активных пациентов (всего)", active_patients, bold=True, fill=_GREEN_FILL_P)
    row += 1

    # Новые пациенты
    new_patients = [float(_safe(r.get("new_patients", 0))) for r in revenue_results]
    _clinic_row("Новых пациентов за месяц", new_patients)
    row += 1

    # Пациентов на клинику в среднем
    pts_per_clinic = [
        round(a / c, 1) if c > 0 else 0.0
        for a, c in zip(active_patients, clinics_by_month)
    ]
    _clinic_row("Ср. пациентов / клиника", pts_per_clinic, fmt=_NUM_FMT_DEC)
    row += 1

    # Для Model A — устройства в парке
    if model_type == "model_a":
        devices = [float(_safe(r.get("devices_in_pool", 0))) for r in revenue_results]
        _clinic_row("Устройств в парке (всего)", devices)
        row += 1

    # Выручка с клиники
    rev_monthly = [_rub(_safe(r.get("total_revenue", 0))) for r in revenue_results]
    rev_per_clinic = [
        round(r / c, 2) if c > 0 else 0.0
        for r, c in zip(rev_monthly, clinics_by_month)
    ]
    _clinic_row("Выручка / клиника (тыс. ₽)", rev_per_clinic, fmt=_NUM_FMT_DEC,
                fill=_YELLOW_FILL_P)
    row += 1


# ──────────────────────────────────────────────
# Лист 8: Бизнес-логика (автогенерированный текст)
# ──────────────────────────────────────────────

def _build_business_logic_sheet(
    wb: Workbook,
    model_type: str,
    num_months: int,
    all_params: Dict,
    revenue_results: List[Dict],
    costs_results: List[Dict],
    cash_flow_results: List[Dict],
    bank_allocation: List[Dict],
    breakeven_result: Dict,
    unit_economics: Dict,
    initial_investment: float,
    clinic_schedule: List[Dict],
) -> None:
    ws = wb.create_sheet("Бизнес-логика")
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 100

    _merge(ws, 1, 1, 1, 2,
           "ОПИСАНИЕ БИЗНЕС-МОДЕЛИ И СТРАТЕГИИ ВЫХОДА НА РЫНОК",
           font=Font(bold=True, size=14, color=_WHITE),
           fill=_HDR_FILL, align="center")
    ws.row_dimensions[1].height = 28

    text_blocks = _generate_business_narrative(
        model_type=model_type,
        num_months=num_months,
        all_params=all_params,
        revenue_results=revenue_results,
        costs_results=costs_results,
        cash_flow_results=cash_flow_results,
        bank_allocation=bank_allocation,
        breakeven_result=breakeven_result,
        unit_economics=unit_economics,
        initial_investment=initial_investment,
        clinic_schedule=clinic_schedule,
    )

    row = 2
    for block in text_blocks:
        title = block.get("title", "")
        content = block.get("content", "")
        if title:
            row += 1
            _merge(ws, row, 1, row, 2,
                   title,
                   font=Font(bold=True, size=12, color=_BLUE),
                   fill=_SECTION_FILL)
            ws.row_dimensions[row].height = 22
            row += 1
        if content:
            cell = ws.merge_cells(f"B{row}:B{row + 5}")
            c = ws.cell(row=row, column=2, value=content)
            c.font = Font(size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = max(18, len(content) // 5)
            row += 6


def _generate_business_narrative(
    model_type: str,
    num_months: int,
    all_params: Dict,
    revenue_results: List[Dict],
    costs_results: List[Dict],
    cash_flow_results: List[Dict],
    bank_allocation: List[Dict],
    breakeven_result: Dict,
    unit_economics: Dict,
    initial_investment: float,
    clinic_schedule: List[Dict],
) -> List[Dict]:
    """
    Генерирует структурированный текст бизнес-логики на основе параметров конфигурации.
    Возвращает список блоков: [{"title": ..., "content": ...}]
    """
    rev_params = all_params.get("revenue", {})
    fix_params = all_params.get("fixed_costs", {})
    var_params = all_params.get("variable_costs", {})
    assumptions = all_params.get("assumptions", {})

    # ── Базовые метрики ──
    total_revenue = sum(_safe(r.get("revenue", 0)) for r in cash_flow_results)
    total_costs = sum(_safe(r.get("total_costs", 0)) for r in cash_flow_results)
    total_cf = sum(_safe(r.get("cash_flow", 0)) for r in cash_flow_results)
    be_month = breakeven_result.get("breakeven_month")
    last_month_rev = _safe(cash_flow_results[-1].get("revenue", 0)) if cash_flow_results else 0
    last_month_patients = int(_safe(revenue_results[-1].get("num_patients", 0))) if revenue_results else 0
    first_month_patients = int(_safe(revenue_results[0].get("num_patients", 0))) if revenue_results else 0

    initial_clinics = int(rev_params.get("num_clinics", 1))
    growth_rate = rev_params.get("growth_rate", 0.0)
    rehab_dur = int(rev_params.get("rehab_duration_months", 3))
    initial_pts_per_clinic = int(rev_params.get("patients_per_clinic_month1", 5))

    # Строим расписание клиник
    all_batches = [{"month_start": 1, "count": initial_clinics}]
    for entry in (clinic_schedule or []):
        ms = int(entry.get("month_start", 2))
        cnt = int(entry.get("count", 0))
        if ms >= 1 and cnt > 0:
            all_batches.append({"month_start": ms, "count": cnt})
    all_batches.sort(key=lambda x: x["month_start"])

    total_bank = sum(_safe(e.get("bank_used", 0)) for e in (bank_allocation or []))
    bank_cats = _bank_category_totals(bank_allocation)

    # ── Блок 1: О продукте ──
    if model_type == "model_a":
        model_desc = (
            "ReFlex работает по модели B2B SaaS: клиника приобретает парк носимых устройств "
            "(ЭМГ+IMU) и подписывается на платформу. Реабилитолог назначает пациентам "
            "сенсоры, мобильное приложение проводит их по программе упражнений, сравнивая "
            "движения с эталоном в реальном времени. Данные агрегируются в отчёт для врача.\n\n"
            f"Устройств на клинику: {int(rev_params.get('devices_per_clinic', 5))} ед. | "
            f"Разовый setup-fee: {rev_params.get('setup_fee', 0):,.0f} ₽/уст. | "
            f"Ежемес. подписка: {rev_params.get('subscription_per_device', 0):,.0f} ₽/уст."
        )
    elif model_type == "model_b":
        commission = rev_params.get("clinic_commission_rate", 0.15)
        model_desc = (
            "ReFlex работает по модели B2B2C: пациент самостоятельно арендует комплект "
            "устройств через клинику. Клиника выступает каналом сбыта и получает комиссию "
            f"({commission * 100:.0f}% от стоимости аренды). Устройства принадлежат ReFlex "
            "и переходят к следующему пациенту после завершения курса реабилитации.\n\n"
            f"Стоимость аренды: {rev_params.get('rental_price_per_patient', 0):,.0f} ₽/мес. | "
            f"Комиссия клиники: {commission * 100:.0f}% | "
            f"Длительность курса: {rehab_dur} мес."
        )
    else:
        model_desc = (
            "ReFlex работает по гибридной модели A+B: клиники с высокой загрузкой подключаются "
            "по модели B2B (парк устройств + подписка), пациенты частных клиник — по модели "
            "B2B2C (аренда). Это позволяет гибко работать с разными типами учреждений.\n\n"
            f"Подписка/уст.: {rev_params.get('subscription_per_device', 0):,.0f} ₽/мес. | "
            f"Аренда/пациент: {rev_params.get('rental_price_per_patient', 0):,.0f} ₽/мес. | "
            f"Горизонт: {num_months} мес."
        )

    blocks = [
        {
            "title": "1. О продукте и бизнес-модели",
            "content": model_desc,
        }
    ]

    # ── Блок 2: GTM-стратегия ──
    gtm_parts = []
    pilot_entry = all_batches[0] if all_batches else {"month_start": 1, "count": initial_clinics}
    pilot_clinics = pilot_entry["count"]
    pilot_end = all_batches[1]["month_start"] - 1 if len(all_batches) > 1 else num_months

    gtm_parts.append(
        f"ЭТАП 1: ПИЛОТ (месяцы 1–{pilot_end})\n"
        f"Запуск с {pilot_clinics} клиник(-ой). Цель — отработать продукт, "
        f"собрать обратную связь от реабилитологов и пациентов, подтвердить "
        f"adherence ≥70%. Ожидаемое число пациентов в М1: "
        f"{initial_clinics * initial_pts_per_clinic} чел. "
        f"(по {initial_pts_per_clinic} чел./клинику)."
    )

    if len(all_batches) > 1:
        for i, batch in enumerate(all_batches[1:], start=2):
            ms = batch["month_start"]
            cnt = batch["count"]
            total_so_far = sum(b["count"] for b in all_batches[:i])
            next_ms = all_batches[i]["month_start"] if i < len(all_batches) else num_months + 1
            end_str = f"месяц {next_ms - 1}" if next_ms <= num_months else f"месяц {num_months} (конец горизонта)"
            gtm_parts.append(
                f"ЭТАП {i}: МАСШТАБИРОВАНИЕ (с месяца {ms})\n"
                f"Подключение {cnt} новых клиник. Итого активных клиник: {total_so_far}. "
                f"Фаза длится до {end_str}."
            )

    if growth_rate > 0:
        gtm_parts.append(
            f"РОСТ ПАЦИЕНТСКОЙ БАЗЫ: ежемесячный прирост новых пациентов "
            f"{growth_rate * 100:.1f}%. При таком темпе к месяцу {num_months} "
            f"ожидается {last_month_patients} активных пациентов "
            f"(старт: {first_month_patients} чел.)."
        )

    blocks.append({
        "title": "2. GTM-стратегия (выход на рынок)",
        "content": "\n\n".join(gtm_parts),
    })

    # ── Блок 3: Динамика клиентской базы ──
    patient_milestones = []
    checkpoints = [num_months // 4, num_months // 2, 3 * num_months // 4, num_months - 1]
    for cp in checkpoints:
        cp = max(0, min(cp, len(revenue_results) - 1))
        pts = int(_safe(revenue_results[cp].get("num_patients", 0)))
        if pts > 0:
            patient_milestones.append(f"М{cp + 1}: {pts} активных пациентов")

    lty = unit_economics.get("ltv", 0)
    cac_val = unit_economics.get("cac", 0)
    ltv_cac = unit_economics.get("ltv_cac_ratio", 0)

    client_text = (
        f"Старт в Месяце 1: {first_month_patients} активных пациентов.\n"
        f"Контрольные точки роста: {' → '.join(patient_milestones)}.\n\n"
    )
    if lty > 0:
        client_text += (
            f"Unit Economics:\n"
            f"  • LTV (ценность пациента за курс): {lty:,.0f} ₽\n"
            f"  • CAC (стоимость привлечения): {cac_val:,.0f} ₽\n"
            f"  • LTV/CAC: {ltv_cac:.2f}x "
            f"({'здоровая модель ✓' if ltv_cac >= 3 else 'требует оптимизации'})\n"
        )
    client_text += (
        f"\nДлительность курса реабилитации: {rehab_dur} мес. — "
        f"ключевой параметр для расчёта когортного удержания и выручки."
    )
    blocks.append({
        "title": "3. Динамика клиентской базы",
        "content": client_text,
    })

    # ── Блок 4: Финансовые ориентиры ──
    dcf_m = [
        round(_rub(_safe(r.get("cash_flow", 0))) / (1.0 + _CAPM_RATE / 4.0) ** (i + 1), 2)
        for i, r in enumerate(cash_flow_results)
    ]
    npv = sum(dcf_m)
    be_str = f"месяц {be_month}" if be_month else "горизонт не достигается (требуется масштабирование)"

    fin_text = (
        f"Горизонт планирования: {num_months} месяцев.\n\n"
        f"Суммарная выручка: {_rub(total_revenue):,.1f} тыс. ₽\n"
        f"Суммарные расходы: {_rub(total_costs):,.1f} тыс. ₽\n"
        f"Суммарный CF: {_rub(total_cf):,.1f} тыс. ₽ "
        f"({'положительный ✓' if total_cf >= 0 else 'отрицательный — требует финансирования'})\n\n"
        f"Точка операционной безубыточности: {be_str}.\n"
        f"NPV (r = {_CAPM_RATE * 100:.1f}%): {npv:,.1f} тыс. ₽\n\n"
        f"Выручка в последнем месяце горизонта: {_rub(last_month_rev):,.1f} тыс. ₽\n"
        f"Число активных пациентов в конце периода: {last_month_patients} чел."
    )
    blocks.append({
        "title": "4. Финансовые ориентиры",
        "content": fin_text,
    })

    # ── Блок 5: Использование гранта ──
    if total_bank > 0:
        cat_label_map = {
            "team_salaries": "Зарплаты команды",
            "infrastructure_fixed": "Инфраструктура",
            "office_rent": "Аренда офиса",
            "legal_services": "Юридические услуги",
            "other_fixed": "Прочие затраты",
            "cogs": "Производство устройств (COGS)",
            "logistics": "Логистика",
            "support": "Поддержка пользователей",
            "infrastructure_variable": "Инфраструктура (переменная)",
            "cac": "Привлечение клиентов (CAC)",
        }
        cat_lines = []
        for cat, lbl in cat_label_map.items():
            val = bank_cats.get(cat, 0.0)
            if val > 0:
                pct = val / total_bank * 100
                cat_lines.append(f"  • {lbl}: {_rub(val):,.1f} тыс. ₽ ({pct:.1f}%)")

        grant_text = (
            f"Объём гранта / начальных инвестиций: {_rub(initial_investment):,.1f} тыс. ₽\n"
            f"Израсходовано в горизонте: {_rub(total_bank):,.1f} тыс. ₽\n\n"
            f"Структура расходования:\n" + "\n".join(cat_lines) if cat_lines else ""
        )
        if not cat_lines:
            grant_text = (
                f"Объём гранта / начальных инвестиций: {_rub(initial_investment):,.1f} тыс. ₽\n"
                "Стратегия расходования настраивается в блоке «Инвестиционный банк» калькулятора."
            )
        blocks.append({
            "title": "5. Использование гранта",
            "content": grant_text,
        })

    # ── Блок 6: Ключевые допущения ──
    churn = assumptions.get("churn_rate", 0.0)
    margin = assumptions.get("desired_margin", 0.0)
    assump_text = (
        f"Ключевые допущения, заложенные в модель:\n"
        f"  • Ежемесячный отток (churn): {churn * 100:.1f}%\n"
        f"  • Целевая маржинальность: {margin * 100:.1f}%\n"
        f"  • Длительность реабилитационного курса: {rehab_dur} мес.\n"
        f"  • Ежемесячный рост числа новых пациентов: {growth_rate * 100:.1f}%\n"
        f"  • Горизонт планирования: {num_months} мес.\n\n"
        "Все допущения являются гипотезами и требуют валидации в ходе пилотного "
        "запуска (см. реестр допущений БП_01)."
    )
    blocks.append({
        "title": "6. Ключевые допущения модели",
        "content": assump_text,
    })

    return blocks


# ──────────────────────────────────────────────
# Главная функция экспорта
# ──────────────────────────────────────────────

def export_to_msp_excel(
    model_type: str,
    all_params: Dict,
    num_months: int,
    revenue_results: List[Dict],
    costs_results: List[Dict],
    cash_flow_results: List[Dict],
    bank_allocation: Optional[List[Dict]],
    breakeven_result: Dict,
    unit_economics: Dict,
    initial_investment: float,
    discount_rate_annual: float,
    clinic_schedule: Optional[List[Dict]] = None,
    filename: Optional[str] = None,
    rnd_results: Optional[List[Dict]] = None,
) -> str:
    """
    Генерирует Excel-файл «БП МСП» для конкурса Студенческий стартап / МСП.

    Args:
        model_type: 'model_a' | 'model_b' | 'model_ab'
        all_params: полный словарь параметров (revenue, fixed_costs, variable_costs, ...)
        num_months: горизонт планирования
        revenue_results: список результатов по месяцам из calculate_revenue_for_months
        costs_results: список результатов по месяцам из calculate_costs_for_months
        cash_flow_results: список результатов по месяцам из calculate_cash_flow_for_months
        bank_allocation: результат calculate_bank_allocation
        breakeven_result: результат calculate_breakeven_month
        unit_economics: результат calculate_unit_economics_from_params
        initial_investment: объём гранта / начальных вложений в рублях
        discount_rate_annual: годовая ставка дисконтирования
        clinic_schedule: список пачек клиник [{month_start, count}]
        filename: имя файла (по умолчанию — автоматическое)

    Returns:
        Путь к созданному файлу.
    """
    if filename is None:
        filename = f"ReFlex_BP_MSP_{model_type}_{date.today().isoformat()}.xlsx"

    if bank_allocation is None:
        bank_allocation = []
    if clinic_schedule is None:
        clinic_schedule = []

    try:
        wb = Workbook()
        # Удаляем стандартный лист
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        # 1. Титульный лист
        _build_title_sheet(
            wb, model_type, all_params, num_months,
            cash_flow_results, breakeven_result, initial_investment,
        )

        # 2. БДДС (по шаблону МСП)
        _build_bdds_sheet(
            wb, model_type, num_months,
            revenue_results, costs_results, cash_flow_results,
            bank_allocation, all_params,
            rnd_results=rnd_results,
        )

        # 3. Инвест (по шаблону МСП)
        _build_invest_sheet(
            wb, num_months,
            cash_flow_results, revenue_results, costs_results,
            bank_allocation, breakeven_result,
            initial_investment, discount_rate_annual,
            rnd_results=rnd_results,
        )

        # 4. График ДП
        _build_cf_chart_sheet(
            wb, num_months, cash_flow_results, bank_allocation,
        )

        # 5. Точка безубыточности
        _build_breakeven_sheet(
            wb, num_months, cash_flow_results, costs_results, breakeven_result,
        )

        # 6. Когорты пациентов
        _build_cohorts_sheet(
            wb, num_months, revenue_results, all_params,
        )

        # 7. Клиники
        _build_clinics_sheet(
            wb, num_months, revenue_results, all_params,
            clinic_schedule, model_type,
        )

        # 8. Бизнес-логика
        _build_business_logic_sheet(
            wb, model_type, num_months, all_params,
            revenue_results, costs_results, cash_flow_results,
            bank_allocation, breakeven_result, unit_economics,
            initial_investment, clinic_schedule,
        )

        wb.save(filename)
        return filename

    except Exception as exc:
        import traceback
        return f"Ошибка при формировании файла: {exc}\n{traceback.format_exc()}"
