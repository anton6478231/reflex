"""
Утилиты для экспорта данных в различные форматы
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional
import json

# Русские метки стандартных статей затрат
_COST_LINE_LABELS: Dict[str, str] = {
    "team_salaries": "Зарплаты команды",
    "infrastructure_fixed": "Инфраструктура (постоянная)",
    "office_rent": "Аренда офиса",
    "legal_services": "Юридические услуги",
    "other_fixed": "Прочие постоянные затраты",
    "cogs": "COGS (производство устройств)",
    "logistics": "Логистика",
    "support": "Поддержка пользователей",
    "infrastructure_variable": "Инфраструктура (переменная)",
    "cac": "CAC (привлечение клиентов)",
}


def _pretty_line_name(name: str) -> str:
    return _COST_LINE_LABELS.get(name, name)


def _add_rnd_sheet(
    wb: Workbook,
    rnd_results: List[Dict],
    header_fill: PatternFill,
    header_font: Font,
) -> None:
    """
    Добавляет лист «R&D расходы» в workbook.

    Строки = категории расходов, столбцы = R&D месяцы.
    Итоговые строка и столбец добавляются автоматически.
    """
    if not rnd_results:
        return

    ws = wb.create_sheet("R&D расходы")

    all_cats = []
    for r in rnd_results:
        for cat in r.get("breakdown", {}).keys():
            if cat not in all_cats:
                all_cats.append(cat)

    num_fmt = "#,##0"
    sum_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    month_labels = [f"R&D {r['month']}" for r in rnd_results]
    total_col = len(month_labels) + 2

    # Заголовок
    ws.cell(1, 1, "R&D РАСХОДЫ ПО МЕСЯЦАМ").font = Font(bold=True, size=13)
    ws.merge_cells(f"A1:{get_column_letter(total_col)}1")
    total_rnd = sum(r["total_costs"] for r in rnd_results)
    ws.cell(2, 1, f"Суммарные расходы R&D: {total_rnd:,.0f} ₽ (из банка инвестиций)".replace(",", " ")).font = Font(italic=True)
    ws.merge_cells(f"A2:{get_column_letter(total_col)}2")

    header_row = 4

    # Заголовки столбцов
    ws.cell(header_row, 1, "Статья расходов").fill = header_fill
    ws.cell(header_row, 1).font = header_font
    ws.cell(header_row, 1).alignment = Alignment(horizontal="left")

    for j, lbl in enumerate(month_labels, start=2):
        c = ws.cell(header_row, j, lbl)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    ws.cell(header_row, total_col, "ИТОГО").fill = header_fill
    ws.cell(header_row, total_col).font = header_font
    ws.cell(header_row, total_col).alignment = Alignment(horizontal="center")

    # Строки данных
    for i, cat in enumerate(all_cats):
        data_row = header_row + 1 + i
        ws.cell(data_row, 1, cat)
        row_total = 0.0
        for j, r in enumerate(rnd_results, start=2):
            val = r["breakdown"].get(cat, 0.0)
            cell = ws.cell(data_row, j, round(val, 2) if val else 0)
            cell.number_format = num_fmt
            cell.alignment = Alignment(horizontal="right")
            row_total += val
        tc = ws.cell(data_row, total_col, round(row_total, 2))
        tc.number_format = num_fmt
        tc.font = Font(bold=True)
        tc.alignment = Alignment(horizontal="right")

    # Итоговая строка
    total_row = header_row + 1 + len(all_cats)
    ws.cell(total_row, 1, "ИТОГО ЗА МЕСЯЦ").font = Font(bold=True)
    ws.cell(total_row, 1).fill = sum_fill

    grand_total = 0.0
    for j, r in enumerate(rnd_results, start=2):
        col_sum = r["total_costs"]
        c = ws.cell(total_row, j, round(col_sum, 2))
        c.number_format = num_fmt
        c.font = Font(bold=True)
        c.fill = sum_fill
        c.alignment = Alignment(horizontal="right")
        grand_total += col_sum

    gt_cell = ws.cell(total_row, total_col, round(grand_total, 2))
    gt_cell.number_format = num_fmt
    gt_cell.font = Font(bold=True, color="1F4E79")
    gt_cell.fill = sum_fill
    gt_cell.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 36
    for j in range(len(month_labels)):
        ws.column_dimensions[get_column_letter(j + 2)].width = 16
    ws.column_dimensions[get_column_letter(total_col)].width = 18


def _add_grant_sheet(
    wb: Workbook,
    bank_allocation: List[Dict],
    header_fill: PatternFill,
    header_font: Font,
) -> None:
    """
    Добавляет лист «Грантовые расходы» в workbook.

    Строки = статьи затрат, покрытые банком.
    Столбцы = месяцы (только те, где банк расходовался).
    Итоговые строка и столбец добавляются автоматически.
    """
    from models.investment_bank import build_grant_matrix

    line_names, month_labels, matrix = build_grant_matrix(bank_allocation)
    if not line_names or not month_labels:
        return

    ws = wb.create_sheet("Грантовые расходы")

    total_bank = sum(e["bank_used"] for e in bank_allocation)

    # ── Заголовок ──
    ws.cell(1, 1, "РАСПРЕДЕЛЕНИЕ ГРАНТОВЫХ/ИНВЕСТИЦИОННЫХ СРЕДСТВ").font = Font(bold=True, size=13)
    ws.merge_cells(f"A1:{get_column_letter(len(month_labels) + 3)}1")
    ws.cell(2, 1, f"Всего потрачено из банка: {total_bank:,.0f} ₽".replace(",", " ")).font = Font(italic=True)
    ws.merge_cells(f"A2:{get_column_letter(len(month_labels) + 3)}2")

    header_row = 4

    # ── Заголовки столбцов ──
    ws.cell(header_row, 1, "Статья затрат").fill = header_fill
    ws.cell(header_row, 1).font = header_font
    ws.cell(header_row, 1).alignment = Alignment(horizontal="left")

    for j, label in enumerate(month_labels, start=2):
        c = ws.cell(header_row, j, label)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    total_col = len(month_labels) + 2
    ws.cell(header_row, total_col, "ИТОГО").fill = header_fill
    ws.cell(header_row, total_col).font = header_font
    ws.cell(header_row, total_col).alignment = Alignment(horizontal="center")

    # ── Строки данных ──
    sum_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    num_fmt = "#,##0"

    for i, (name, row_vals) in enumerate(zip(line_names, matrix)):
        data_row = header_row + 1 + i
        ws.cell(data_row, 1, _pretty_line_name(name))
        row_total = 0.0
        for j, val in enumerate(row_vals, start=2):
            cell = ws.cell(data_row, j, round(val, 2) if val else 0)
            cell.number_format = num_fmt
            cell.alignment = Alignment(horizontal="right")
            row_total += val
        total_cell = ws.cell(data_row, total_col, round(row_total, 2))
        total_cell.number_format = num_fmt
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal="right")

    # ── Итоговая строка (по месяцам) ──
    total_row = header_row + 1 + len(line_names)
    ws.cell(total_row, 1, "ИТОГО ЗА МЕСЯЦ").font = Font(bold=True)
    ws.cell(total_row, 1).fill = sum_fill

    grand_total = 0.0
    for j in range(len(month_labels)):
        col_sum = sum(matrix[i][j] for i in range(len(line_names)))
        c = ws.cell(total_row, j + 2, round(col_sum, 2))
        c.number_format = num_fmt
        c.font = Font(bold=True)
        c.fill = sum_fill
        c.alignment = Alignment(horizontal="right")
        grand_total += col_sum

    gt_cell = ws.cell(total_row, total_col, round(grand_total, 2))
    gt_cell.number_format = num_fmt
    gt_cell.font = Font(bold=True, color="1F4E79")
    gt_cell.fill = sum_fill
    gt_cell.alignment = Alignment(horizontal="right")

    # ── Ширина колонок ──
    ws.column_dimensions["A"].width = 36
    for j in range(len(month_labels)):
        ws.column_dimensions[get_column_letter(j + 2)].width = 16
    ws.column_dimensions[get_column_letter(total_col)].width = 18


def export_to_excel_with_formulas(
    model_type: str,
    all_params: Dict,
    num_months: int,
    filename: str = "reflex_fem_with_formulas.xlsx",
    bank_allocation: Optional[List[Dict]] = None,
    rnd_results: Optional[List[Dict]] = None,
) -> str:
    """
    Экспорт ФЭМ в Excel с формулами и логикой
    Файл будет полностью интерактивным - можно менять параметры и все пересчитается
    
    Returns:
        Путь к созданному файлу
    """
    try:
        wb = Workbook()
        
        # Удаляем дефолтный лист
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Стили
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        param_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        param_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ============ Лист 1: Параметры ============
        ws_params = wb.create_sheet("Параметры")
        
        row = 1
        ws_params.cell(row, 1, "ФИНАНСОВО-ЭКОНОМИЧЕСКАЯ МОДЕЛЬ REFLEX").font = Font(bold=True, size=14)
        ws_params.merge_cells(f'A{row}:C{row}')
        row += 1
        
        ws_params.cell(row, 1, f"Модель: {model_type}").font = Font(italic=True)
        row += 1
        ws_params.cell(row, 1, f"Горизонт: {num_months} месяцев").font = Font(italic=True)
        row += 2
        
        # Заголовок
        ws_params.cell(row, 1, "Параметр").fill = header_fill
        ws_params.cell(row, 1).font = header_font
        ws_params.cell(row, 2, "Значение").fill = header_fill
        ws_params.cell(row, 2).font = header_font
        ws_params.cell(row, 3, "Единица").fill = header_fill
        ws_params.cell(row, 3).font = header_font
        row += 1
        
        # Revenue параметры
        ws_params.cell(row, 1, "REVENUE PARAMETERS").fill = param_fill
        ws_params.cell(row, 1).font = param_font
        ws_params.merge_cells(f'A{row}:C{row}')
        row += 1
        
        param_rows = {}  # Для ссылок на параметры
        
        clinic_schedule = []
        for key, value in all_params['revenue'].items():
            if key == 'clinic_schedule':
                clinic_schedule = value or []
                continue  # выводим отдельной секцией
            ws_params.cell(row, 1, key)
            cell = ws_params.cell(row, 2, value)
            cell.number_format = '#,##0.00' if isinstance(value, float) else '#,##0'
            ws_params.cell(row, 3, _get_unit(key))
            param_rows[f'revenue_{key}'] = f'Параметры!B{row}'
            row += 1

        row += 1

        # Расписание клиник (только для model_a при наличии)
        if model_type == 'model_a':
            ws_params.cell(row, 1, "РАСПИСАНИЕ КЛИНИК (clinic_schedule)").fill = param_fill
            ws_params.cell(row, 1).font = param_font
            ws_params.merge_cells(f'A{row}:C{row}')
            row += 1
            ws_params.cell(row, 1, "Пачка #")
            ws_params.cell(row, 2, "Месяц старта")
            ws_params.cell(row, 3, "Клиник в пачке")
            for col in range(1, 4):
                ws_params.cell(row, col).font = Font(bold=True)
            row += 1
            if clinic_schedule:
                for i, entry in enumerate(clinic_schedule, start=1):
                    ws_params.cell(row, 1, i)
                    ws_params.cell(row, 2, entry.get('month_start', '?'))
                    ws_params.cell(row, 3, entry.get('count', '?'))
                    row += 1
            else:
                ws_params.cell(row, 1, "(нет дополнительных пачек — все клиники стартуют в M1)")
                ws_params.merge_cells(f'A{row}:C{row}')
                row += 1
        
        row += 1
        
        # Fixed Costs
        ws_params.cell(row, 1, "FIXED COSTS (ежемесячные)").fill = param_fill
        ws_params.cell(row, 1).font = param_font
        ws_params.merge_cells(f'A{row}:C{row}')
        row += 1
        
        for key, value in all_params['fixed_costs'].items():
            ws_params.cell(row, 1, key)
            cell = ws_params.cell(row, 2, value)
            cell.number_format = '#,##0'
            ws_params.cell(row, 3, "₽/мес")
            param_rows[f'fixed_{key}'] = f'Параметры!B{row}'
            row += 1
        
        row += 1
        
        # Variable Costs
        ws_params.cell(row, 1, "VARIABLE COSTS").fill = param_fill
        ws_params.cell(row, 1).font = param_font
        ws_params.merge_cells(f'A{row}:C{row}')
        row += 1
        
        for key, value in all_params['variable_costs'].items():
            if not key.startswith('custom_'):  # Пропускаем, обработаем отдельно
                ws_params.cell(row, 1, key)
                cell = ws_params.cell(row, 2, value)
                cell.number_format = '#,##0'
                ws_params.cell(row, 3, _get_unit(key))
                param_rows[f'variable_{key}'] = f'Параметры!B{row}'
                row += 1
        
        row += 1
        
        # Кастомные Fixed Costs
        if all_params.get('custom_fixed_costs'):
            ws_params.cell(row, 1, "CUSTOM FIXED COSTS").fill = param_fill
            ws_params.cell(row, 1).font = param_font
            ws_params.merge_cells(f'A{row}:C{row}')
            row += 1
            
            for name, data in all_params['custom_fixed_costs'].items():
                ws_params.cell(row, 1, name)
                cell = ws_params.cell(row, 2, data['value'])
                cell.number_format = '#,##0'
                ws_params.cell(row, 3, data['type'])
                param_rows[f'custom_fixed_{name}'] = f'Параметры!B{row}'
                row += 1
            
            row += 1
        
        # Кастомные Variable Costs
        if all_params.get('custom_variable_costs'):
            ws_params.cell(row, 1, "CUSTOM VARIABLE COSTS").fill = param_fill
            ws_params.cell(row, 1).font = param_font
            ws_params.merge_cells(f'A{row}:C{row}')
            row += 1
            
            for name, data in all_params['custom_variable_costs'].items():
                ws_params.cell(row, 1, name)
                cell = ws_params.cell(row, 2, data['value'])
                cell.number_format = '#,##0'
                ws_params.cell(row, 3, data['type'])
                param_rows[f'custom_var_{name}'] = f'Параметры!B{row}'
                row += 1
        
        # Устанавливаем ширину колонок
        ws_params.column_dimensions['A'].width = 35
        ws_params.column_dimensions['B'].width = 15
        ws_params.column_dimensions['C'].width = 15
        
        # ============ Лист 2: Расчеты по месяцам ============
        ws_calc = wb.create_sheet("Расчеты")
        
        row = 1
        ws_calc.cell(row, 1, "РАСЧЕТЫ ПО МЕСЯЦАМ").font = Font(bold=True, size=14)
        ws_calc.merge_cells(f'A{row}:H{row}')
        row += 2
        
        # Заголовки
        headers = ["Месяц", "Revenue", "Fixed Costs", "Variable Costs", "Total Costs", "Cash Flow", "Cumulative CF", ""]
        for col, header in enumerate(headers, 1):
            cell = ws_calc.cell(row, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Формулы для расчетов
        start_row = row
        
        for month in range(1, num_months + 1):
            ws_calc.cell(row, 1, month).alignment = Alignment(horizontal='center')
            
            # Revenue формула (зависит от модели)
            revenue_formula = _get_revenue_formula(model_type, month, param_rows, num_months)
            ws_calc.cell(row, 2, revenue_formula)
            ws_calc.cell(row, 2).number_format = '#,##0'
            
            # Fixed Costs - сумма всех fixed costs
            fixed_refs = [param_rows[k] for k in param_rows if k.startswith('fixed_') and not k.startswith('fixed_custom_')]
            
            # Добавляем кастомные fixed costs (с учетом типа)
            custom_fixed_parts = []
            if all_params.get('custom_fixed_costs'):
                for name, data in all_params['custom_fixed_costs'].items():
                    ref = param_rows.get(f'custom_fixed_{name}')
                    if ref:
                        if data['type'] == "Единоразовая (месяц 1)" and month != 1:
                            continue  # Пропускаем для месяцев 2+
                        custom_fixed_parts.append(ref)
            
            all_fixed_refs = fixed_refs + custom_fixed_parts
            fixed_formula = f"={'+'.join(all_fixed_refs)}" if all_fixed_refs else "=0"
            ws_calc.cell(row, 3, fixed_formula)
            ws_calc.cell(row, 3).number_format = '#,##0'
            
            # Variable Costs - формула зависит от модели
            var_formula = _get_variable_costs_formula(model_type, month, param_rows, row, start_row, all_params)
            ws_calc.cell(row, 4, var_formula)
            ws_calc.cell(row, 4).number_format = '#,##0'
            
            # Total Costs = Fixed + Variable
            ws_calc.cell(row, 5, f"=C{row}+D{row}")
            ws_calc.cell(row, 5).number_format = '#,##0'
            
            # Cash Flow = Revenue - Total Costs
            ws_calc.cell(row, 6, f"=B{row}-E{row}")
            ws_calc.cell(row, 6).number_format = '#,##0'
            
            # Cumulative CF
            if month == 1:
                ws_calc.cell(row, 7, f"=F{row}")
            else:
                ws_calc.cell(row, 7, f"=G{row-1}+F{row}")
            ws_calc.cell(row, 7).number_format = '#,##0'
            
            row += 1
        
        # Итого
        row += 1
        ws_calc.cell(row, 1, "ИТОГО:").font = Font(bold=True)
        for col in range(2, 8):
            if col != 7:  # Кроме Cumulative CF
                ws_calc.cell(row, col, f"=SUM({get_column_letter(col)}{start_row}:{get_column_letter(col)}{row-2})")
                ws_calc.cell(row, col).font = Font(bold=True)
                ws_calc.cell(row, col).number_format = '#,##0'
        
        # Устанавливаем ширину колонок
        for col in range(1, 9):
            ws_calc.column_dimensions[get_column_letter(col)].width = 15
        
        # ============ Лист 3: KPI метрики ============
        ws_kpi = wb.create_sheet("KPI")
        
        row = 1
        ws_kpi.cell(row, 1, "КЛЮЧЕВЫЕ МЕТРИКИ").font = Font(bold=True, size=14)
        ws_kpi.merge_cells(f'A{row}:C{row}')
        row += 2
        
        # Заголовки
        ws_kpi.cell(row, 1, "Метрика").fill = header_fill
        ws_kpi.cell(row, 1).font = header_font
        ws_kpi.cell(row, 2, "Значение").fill = header_fill
        ws_kpi.cell(row, 2).font = header_font
        ws_kpi.cell(row, 3, "Комментарий").fill = header_fill
        ws_kpi.cell(row, 3).font = header_font
        row += 1
        
        # Total Revenue
        ws_kpi.cell(row, 1, f"Total Revenue ({num_months} мес)")
        ws_kpi.cell(row, 2, f"=Расчеты!B{start_row + num_months + 1}")
        ws_kpi.cell(row, 2).number_format = '#,##0 ₽'
        ws_kpi.cell(row, 3, "Общая выручка за период")
        row += 1
        
        # Total Costs
        ws_kpi.cell(row, 1, f"Total Costs ({num_months} мес)")
        ws_kpi.cell(row, 2, f"=Расчеты!E{start_row + num_months + 1}")
        ws_kpi.cell(row, 2).number_format = '#,##0 ₽'
        ws_kpi.cell(row, 3, "Общие затраты за период")
        row += 1
        
        # Net Cash Flow
        ws_kpi.cell(row, 1, f"Net Cash Flow ({num_months} мес)")
        ws_kpi.cell(row, 2, f"=Расчеты!F{start_row + num_months + 1}")
        ws_kpi.cell(row, 2).number_format = '#,##0 ₽'
        ws_kpi.cell(row, 3, "Прибыль/Убыток за период")
        row += 1
        
        # Average monthly CF
        ws_kpi.cell(row, 1, "Средний месячный CF")
        ws_kpi.cell(row, 2, f"=B{row}/({num_months})")
        ws_kpi.cell(row, 2).number_format = '#,##0 ₽'
        ws_kpi.cell(row, 3, "Среднемесячный денежный поток")
        row += 1
        
        # Margin
        ws_kpi.cell(row, 1, "Маржинальность (%)")
        ws_kpi.cell(row, 2, f"=(B{row-1}/B{row-4})*100")
        ws_kpi.cell(row, 2).number_format = '0.0"%"'
        ws_kpi.cell(row, 3, "Процент прибыли от выручки")
        row += 1
        
        # Unit Economics (для model_b)
        if model_type in ['model_b', 'model_ab']:
            row += 1
            ws_kpi.cell(row, 1, "UNIT ECONOMICS").fill = param_fill
            ws_kpi.cell(row, 1).font = param_font
            ws_kpi.merge_cells(f'A{row}:C{row}')
            row += 1
            
            # LTV
            ws_kpi.cell(row, 1, "LTV (Lifetime Value)")
            rental_price_ref = param_rows.get('revenue_rental_price', 'Параметры!B10')
            avg_duration_ref = param_rows.get('revenue_avg_rental_duration', 'Параметры!B11')
            commission_ref = param_rows.get('revenue_clinic_commission_rate', 'Параметры!B12')
            
            ws_kpi.cell(row, 2, f"={rental_price_ref}*(1-{commission_ref})*{avg_duration_ref}")
            ws_kpi.cell(row, 2).number_format = '#,##0 ₽'
            ws_kpi.cell(row, 3, "Ценность клиента за время аренды")
            row += 1
        
        # Устанавливаем ширину колонок
        ws_kpi.column_dimensions['A'].width = 30
        ws_kpi.column_dimensions['B'].width = 20
        ws_kpi.column_dimensions['C'].width = 40
        
        # ============ Лист 4: Инструкция ============
        ws_help = wb.create_sheet("Инструкция")
        
        row = 1
        ws_help.cell(row, 1, "КАК ИСПОЛЬЗОВАТЬ ЭТОТ ФАЙЛ").font = Font(bold=True, size=14, color="4472C4")
        ws_help.merge_cells(f'A{row}:D{row}')
        row += 2
        
        instructions = [
            ("1. Изменение параметров:", "Перейдите на лист 'Параметры' и измените нужные значения в колонке B"),
            ("", "Все расчеты обновятся автоматически!"),
            ("", ""),
            ("2. Просмотр расчетов:", "Лист 'Расчеты' показывает месячную динамику"),
            ("", "Формулы автоматически пересчитываются при изменении параметров"),
            ("", ""),
            ("3. KPI метрики:", "Лист 'KPI' показывает ключевые показатели"),
            ("", "Все значения связаны формулами с параметрами и расчетами"),
            ("", ""),
            ("4. Горизонт планирования:", f"Текущий файл рассчитан на {num_months} месяцев"),
            ("", "Для изменения горизонта используйте калькулятор ReFlex"),
            ("", ""),
            ("5. Модель:", f"Текущая модель: {model_type}"),
            ("", "Model A (B2B) - клиника покупает устройства"),
            ("", "Model B (B2B2C) - пациент арендует через клинику"),
            ("", "Model A+B (Гибрид) - клиника покупает и сдает в аренду"),
        ]
        
        for instruction in instructions:
            if instruction[0]:
                ws_help.cell(row, 1, instruction[0]).font = Font(bold=True)
            ws_help.cell(row, 2, instruction[1])
            ws_help.merge_cells(f'B{row}:D{row}')
            row += 1
        
        ws_help.column_dimensions['A'].width = 25
        ws_help.column_dimensions['B'].width = 60

        # ── Лист «Грантовые расходы» (если передан bank_allocation) ──
        if bank_allocation:
            _add_grant_sheet(wb, bank_allocation, header_fill, header_font)

        # ── Лист «R&D расходы» (если R&D фаза включена) ──
        if rnd_results:
            _add_rnd_sheet(wb, rnd_results, header_fill, header_font)

        # Сохраняем
        wb.save(filename)
        return filename
    
    except Exception as e:
        return f"Ошибка при экспорте: {str(e)}"


def _get_unit(param_name: str) -> str:
    """Возвращает единицу измерения для параметра"""
    units = {
        'num_clinics': 'шт',
        'devices_per_clinic': 'шт',
        'setup_fee': '₽',
        'subscription_per_device': '₽/мес',
        'patients_per_clinic_month1': 'чел',
        'growth_rate': '%',
        'rental_price': '₽/мес',
        'avg_rental_duration': 'мес',
        'clinic_commission_rate': '%',
        'cogs_per_device': '₽',
        'logistics_per_patient': '₽',
        'support_per_patient_per_month': '₽/мес',
        'cac_clinic': '₽',
        'cac_patient': '₽',
        'infrastructure_per_user': '₽/мес',
    }
    return units.get(param_name, '')


def _get_revenue_formula(model_type: str, month: int, param_rows: Dict, num_months: int) -> str:
    """Генерирует формулу для расчета revenue в зависимости от модели"""
    if model_type == 'model_a':
        # Model A: Setup Fee (только месяц 1) + Subscription
        num_clinics = param_rows['revenue_num_clinics']
        devices = param_rows['revenue_devices_per_clinic']
        setup_fee = param_rows['revenue_setup_fee']
        subscription = param_rows['revenue_subscription_per_device']
        
        if month == 1:
            return f"={num_clinics}*{devices}*{setup_fee}+{num_clinics}*{devices}*{subscription}"
        else:
            return f"={num_clinics}*{devices}*{subscription}"
    
    elif model_type == 'model_b':
        # Model B: выручка считается по активным пациентам (кохорты за период реабилитации)
        num_clinics = param_rows['revenue_num_clinics']
        patients_m1 = param_rows['revenue_patients_per_clinic_month1']
        growth = param_rows['revenue_growth_rate']
        price = param_rows['revenue_rental_price']
        commission = param_rows['revenue_clinic_commission_rate']
        duration = param_rows.get('revenue_rehab_duration_months', param_rows.get('revenue_avg_rental_duration', '1'))
        patients_formula = _get_active_patients_formula(num_clinics, patients_m1, growth, duration, month)
        return f"={patients_formula}*{price}*(1-{commission})"
    
    elif model_type == 'model_ab':
        # Model A+B: Setup (м1) + Subscription + Rental
        num_clinics = param_rows['revenue_num_clinics']
        devices = param_rows['revenue_devices_per_clinic']
        setup_fee = param_rows['revenue_setup_fee']
        subscription = param_rows['revenue_subscription_per_device']
        patients_m1 = param_rows['revenue_patients_per_clinic_month1']
        growth = param_rows['revenue_growth_rate']
        price = param_rows['revenue_rental_price']
        commission = param_rows['revenue_clinic_commission_rate']
        
        # Model A часть
        if month == 1:
            model_a_part = f"{num_clinics}*{devices}*{setup_fee}+{num_clinics}*{devices}*{subscription}"
        else:
            model_a_part = f"{num_clinics}*{devices}*{subscription}"
        
        # Model B часть (активные пациенты с учетом срока реабилитации)
        duration = param_rows.get('revenue_rehab_duration_months', param_rows.get('revenue_avg_rental_duration', '1'))
        patients_formula = _get_active_patients_formula(num_clinics, patients_m1, growth, duration, month)
        model_b_part = f"{patients_formula}*{price}*(1-{commission})"
        
        return f"={model_a_part}+{model_b_part}"
    
    return "=0"


def _get_variable_costs_formula(model_type: str, month: int, param_rows: Dict, current_row: int, start_row: int, all_params: Dict = None) -> str:
    """Генерирует формулу для переменных затрат"""
    base_formula_parts = []
    
    if model_type == 'model_a':
        # COGS только в месяц 1
        num_clinics = param_rows['revenue_num_clinics']
        devices = param_rows['revenue_devices_per_clinic']
        cogs = param_rows['variable_cogs_per_device']
        
        if month == 1:
            base_formula_parts.append(f"{num_clinics}*{devices}*{cogs}")
    
    elif model_type in ['model_b', 'model_ab']:
        # Переменные затраты на пациентов
        num_clinics = param_rows['revenue_num_clinics']
        patients_m1 = param_rows['revenue_patients_per_clinic_month1']
        growth = param_rows['revenue_growth_rate']
        logistics = param_rows['variable_logistics_per_patient']
        support = param_rows['variable_support_per_patient_per_month']
        infra = param_rows['variable_infrastructure_per_user']
        
        duration = param_rows.get('revenue_rehab_duration_months', param_rows.get('revenue_avg_rental_duration', '1'))
        patients_formula = _get_active_patients_formula(num_clinics, patients_m1, growth, duration, month)
        
        if model_type == 'model_ab' and month == 1:
            # Добавляем COGS для model_ab в месяц 1
            devices = param_rows['revenue_devices_per_clinic']
            cogs = param_rows['variable_cogs_per_device']
            base_formula_parts.append(f"{num_clinics}*{devices}*{cogs}")
        
        base_formula_parts.append(f"{patients_formula}*({logistics}+{support}+{infra})")
    
    # Добавляем кастомные variable costs
    if all_params and all_params.get('custom_variable_costs'):
        num_clinics = param_rows.get('revenue_num_clinics', 'Параметры!B7')
        devices = param_rows.get('revenue_devices_per_clinic', 'Параметры!B8')
        patients_m1 = param_rows.get('revenue_patients_per_clinic_month1', 'Параметры!B9')
        growth = param_rows.get('revenue_growth_rate', 'Параметры!B10')
        
        for name, data in all_params['custom_variable_costs'].items():
            ref = param_rows.get(f'custom_var_{name}')
            if not ref:
                continue
            
            var_type = data['type']
            
            if var_type == "На устройство (разово)" and month == 1:
                base_formula_parts.append(f"{num_clinics}*{devices}*{ref}")
            elif var_type == "На пациента (разово)" and month == 1:
                patients = _get_active_patients_formula(num_clinics, patients_m1, growth, 1, month)
                base_formula_parts.append(f"{patients}*{ref}")
            elif var_type == "На пациента/месяц":
                duration = param_rows.get('revenue_rehab_duration_months', param_rows.get('revenue_avg_rental_duration', '1'))
                patients = _get_active_patients_formula(num_clinics, patients_m1, growth, duration, month)
                base_formula_parts.append(f"{patients}*{ref}")
            elif var_type == "На клинику (разово)" and month == 1:
                base_formula_parts.append(f"{num_clinics}*{ref}")
    
    if base_formula_parts:
        return f"={'+'.join(base_formula_parts)}"
    return "=0"


def _get_active_patients_formula(
    num_clinics_ref: str,
    patients_m1_ref: str,
    growth_ref: str,
    duration_ref: str,
    month: int,
) -> str:
    """Excel formula for active patients as sum of last D new-patient cohorts."""
    terms = []
    for idx in range(0, month):
        cohort_month = month - idx
        terms.append(
            "IF("
            f"{idx}<{duration_ref},"
            f"ROUND({num_clinics_ref}*{patients_m1_ref}*POWER(1+{growth_ref},{cohort_month - 1}),0),"
            "0)"
        )
    return f"({' + '.join(terms)})"


def export_to_excel(
    cash_flow_results: List[Dict],
    revenue_results: List[Dict],
    costs_results: List[Dict],
    unit_economics: Dict,
    all_params: Dict,
    filename: str = "reflex_calculator_results.xlsx",
    bank_allocation: Optional[List[Dict]] = None,
    rnd_results: Optional[List[Dict]] = None,
) -> str:
    """
    Простой экспорт результатов в Excel (без формул, только данные)
    
    Returns:
        Путь к созданному файлу
    """
    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Лист 1: Параметры
            params_data = []
            
            # Revenue params
            for key, value in all_params['revenue'].items():
                params_data.append({
                    'Категория': 'Revenue',
                    'Параметр': key,
                    'Значение': value
                })
            
            # Fixed costs
            for key, value in all_params['fixed_costs'].items():
                params_data.append({
                    'Категория': 'Fixed Costs',
                    'Параметр': key,
                    'Значение': value
                })
            
            # Кастомные Fixed costs
            if all_params.get('custom_fixed_costs'):
                for name, data in all_params['custom_fixed_costs'].items():
                    params_data.append({
                        'Категория': f'Custom Fixed ({data["type"]})',
                        'Параметр': name,
                        'Значение': data['value']
                    })
            
            # Variable costs
            for key, value in all_params['variable_costs'].items():
                params_data.append({
                    'Категория': 'Variable Costs',
                    'Параметр': key,
                    'Значение': value
                })
            
            # Кастомные Variable costs
            if all_params.get('custom_variable_costs'):
                for name, data in all_params['custom_variable_costs'].items():
                    params_data.append({
                        'Категория': f'Custom Variable ({data["type"]})',
                        'Параметр': name,
                        'Значение': data['value']
                    })
            
            df_params = pd.DataFrame(params_data)
            df_params.to_excel(writer, sheet_name='Параметры', index=False)
            
            # Лист 2: Расчеты по месяцам
            cf_data = []
            for i, cf in enumerate(cash_flow_results):
                revenue = revenue_results[i]
                costs = costs_results[i]
                
                cf_data.append({
                    'Месяц': cf['month'],
                    'Revenue': cf['revenue'],
                    'Fixed Costs': cf['fixed_costs'],
                    'Variable Costs': cf['variable_costs'],
                    'Total Costs': cf['total_costs'],
                    'Cash Flow': cf['cash_flow'],
                    'Cumulative CF': cf['cumulative_cash_flow']
                })
            
            df_cf = pd.DataFrame(cf_data)
            df_cf.to_excel(writer, sheet_name='Cash Flow', index=False)
            
            # Лист 3: Unit Economics (если есть)
            if unit_economics and unit_economics.get('ltv', 0) > 0:
                ue_data = [{
                    'Метрика': key,
                    'Значение': value
                } for key, value in unit_economics.items()]
                
                df_ue = pd.DataFrame(ue_data)
                df_ue.to_excel(writer, sheet_name='Unit Economics', index=False)

            # Лист 4: Грантовые расходы (если передан bank_allocation)
            if bank_allocation:
                from models.investment_bank import build_grant_matrix
                line_names, month_labels, matrix = build_grant_matrix(bank_allocation)
                if line_names and month_labels:
                    pretty_rows = [_pretty_line_name(n) for n in line_names]
                    df_grant = pd.DataFrame(matrix, index=pretty_rows, columns=month_labels)
                    df_grant["ИТОГО"] = df_grant.sum(axis=1)
                    total_row = df_grant.sum(axis=0)
                    total_row.name = "ИТОГО"
                    df_grant = pd.concat([df_grant, total_row.to_frame().T])
                    df_grant.to_excel(writer, sheet_name='Грантовые расходы')

            # Лист 5: R&D расходы (если R&D фаза включена)
            if rnd_results:
                rnd_rows = []
                all_cats = sorted(
                    {cat for r in rnd_results for cat in r.get("breakdown", {}).keys()}
                )
                for r in rnd_results:
                    row: Dict = {"Период": f"R&D {r['month']}"}
                    for cat in all_cats:
                        row[cat] = r["breakdown"].get(cat, 0.0)
                    row["Итого расходов"] = r["total_costs"]
                    row["Cash Flow"] = r["cash_flow"]
                    row["Cumulative CF"] = r["cumulative_cash_flow"]
                    rnd_rows.append(row)
                if rnd_rows:
                    df_rnd = pd.DataFrame(rnd_rows)
                    df_rnd.to_excel(writer, sheet_name="R&D расходы", index=False)

        return filename
    
    except Exception as e:
        return f"Ошибка при экспорте: {str(e)}"


def _fmt_rub(value: float) -> str:
    if value is None:
        return "—"
    if value == float("inf") or value == float("-inf"):
        return "∞" if value > 0 else "-∞"
    return f"{value:,.0f}".replace(",", " ") + " ₽"


def _fmt_pct(value: float) -> str:
    if value is None:
        return "—"
    return f"{value * 100:.1f}%"


def build_bp04_fem_snapshot_markdown(
    model_type: str,
    model_display_name: str,
    scenario_label: str,
    num_months: int,
    all_params: Dict,
    cash_flow_results: List[Dict],
    revenue_results: List[Dict],
    unit_economics: Dict,
    breakeven_result: Dict,
    min_rental_price: float,
    rnd_results: Optional[List[Dict]] = None,
) -> str:
    """
    Снимок настроек и результатов калькулятора в Markdown для передачи в LLM
    и последующего внесения в БП_04_Финансово_экономическая_модель_ReFlex.md
    """
    from datetime import date

    total_rev = sum(cf["revenue"] for cf in cash_flow_results)
    total_costs = sum(cf["total_costs"] for cf in cash_flow_results)
    net_cf = sum(cf["cash_flow"] for cf in cash_flow_results)
    be_text = (
        f"месяц {breakeven_result['breakeven_month']}"
        if breakeven_result.get("reached")
        else "не достигается в проекции"
    )

    lines: List[str] = []
    lines.append("---")
    lines.append("source: reflex-calculator")
    lines.append("target_document: Общая информация о проекте ReFlex/БП_04_Финансово_экономическая_модель_ReFlex.md")
    lines.append(f"generated_date: {date.today().isoformat()}")
    lines.append(f"model_type: {model_type}")
    lines.append(f"model_display: {model_display_name}")
    lines.append(f"scenario: {scenario_label}")
    lines.append(f"horizon_months: {num_months}")
    lines.append("---")
    lines.append("")
    lines.append("# Снимок ФЭМ из калькулятора ReFlex (для БП_04)")
    lines.append("")
    lines.append(
        "Этот файл сгенерирован калькулятором `reflex-calculator`. "
        "Передай его ассистенту в Cursor с запросом обновить разделы БП_04 по этим данным."
    )
    lines.append("")
    lines.append("## Инструкция для ассистента (LLM)")
    lines.append("")
    lines.append(
        "1. Открой целевой документ: `Общая информация о проекте ReFlex/БП_04_Финансово_экономическая_модель_ReFlex.md`."
    )
    lines.append(
        "2. Замени или дополни блок «(Заполнить после прохождения Этапов 1–5 roadmap.)» "
        "структурированным содержанием из разделов ниже (параметры, допущения, помесячный CF, KPI)."
    )
    lines.append(
        "3. Явно пометь validated vs гипотеза по правилам проекта; ссылайся на калькулятор как источник сценария."
    )
    lines.append(
        "4. Сохрани связь с roadmap: Этап 6 — Финансовая модель (`Разработка_бизнес_плана_ReFlex_пошагово.md`)."
    )
    lines.append("")
    lines.append("## Контекст сценария")
    lines.append("")
    lines.append(f"| Поле | Значение |")
    lines.append(f"|------|----------|")
    lines.append(f"| Модель | {model_display_name} (`{model_type}`) |")
    lines.append(f"| Сценарий / режим | {scenario_label} |")
    lines.append(f"| Горизонт | {num_months} мес. |")
    lines.append("")
    lines.append("## Сводные KPI за период")
    lines.append("")
    lines.append(f"| Метрика | Значение |")
    lines.append(f"|---------|----------|")
    lines.append(f"| Total Revenue | {_fmt_rub(total_rev)} |")
    lines.append(f"| Total Costs | {_fmt_rub(total_costs)} |")
    lines.append(f"| Net Cash Flow | {_fmt_rub(net_cf)} |")
    lines.append(f"| Breakeven (проекция) | {be_text} |")
    lines.append(
        f"| Мин. цена аренды (ориентир) | {_fmt_rub(min_rental_price) if min_rental_price != float('inf') else '— (не применимо к модели)'} |"
    )
    if unit_economics and unit_economics.get("ltv"):
        lines.append(f"| LTV | {_fmt_rub(unit_economics.get('ltv', 0))} |")
        lines.append(f"| CAC (пациент) | {_fmt_rub(unit_economics.get('cac', 0))} |")
        ratio = unit_economics.get("ltv_cac_ratio", 0)
        lines.append(
            f"| LTV/CAC | {'∞' if ratio == float('inf') else f'{ratio:.2f}'} |"
        )
    lines.append("")
    lines.append("## Параметры (вводные)")
    lines.append("")
    lines.append("### Revenue")
    lines.append("")
    lines.append("| Параметр | Значение |")
    lines.append("|----------|----------|")
    for k, v in sorted(all_params.get("revenue", {}).items()):
        if isinstance(v, float) and 0 <= v <= 1 and (
            "rate" in k or "growth" in k or "commission" in k
        ):
            lines.append(f"| `{k}` | {_fmt_pct(v)} |")
        else:
            lines.append(f"| `{k}` | {v} |")
    lines.append("")
    lines.append("### Fixed costs (базовые)")
    lines.append("")
    lines.append("| Параметр | Значение (₽/мес) |")
    lines.append("|----------|------------------|")
    for k, v in sorted(all_params.get("fixed_costs", {}).items()):
        if not str(k).startswith("custom_"):
            lines.append(f"| `{k}` | {v:,.0f} |".replace(",", " "))
    lines.append("")
    lines.append("### Variable costs (базовые)")
    lines.append("")
    lines.append("| Параметр | Значение |")
    lines.append("|----------|----------|")
    for k, v in sorted(all_params.get("variable_costs", {}).items()):
        if not str(k).startswith("custom_"):
            lines.append(f"| `{k}` | {v} |")
    lines.append("")
    lines.append("### Assumptions")
    lines.append("")
    lines.append("| Параметр | Значение |")
    lines.append("|----------|----------|")
    for k, v in sorted(all_params.get("assumptions", {}).items()):
        if isinstance(v, float) and 0 <= v <= 1 and (
            "margin" in k or "rate" in k or "utilization" in k or "churn" in k
        ):
            lines.append(f"| `{k}` | {_fmt_pct(v)} |")
        else:
            lines.append(f"| `{k}` | {v} |")
    lines.append("")

    cf = all_params.get("custom_fixed_costs") or {}
    cv = all_params.get("custom_variable_costs") or {}
    if cf:
        lines.append("### Кастомные Fixed costs")
        lines.append("")
        lines.append("| Название | Сумма | Тип |")
        lines.append("|----------|-------|-----|")
        for name, data in cf.items():
            lines.append(f"| {name} | {data.get('value', 0):,.0f} ₽ | {data.get('type', '')} |".replace(",", " "))
        lines.append("")
    if cv:
        lines.append("### Кастомные Variable costs")
        lines.append("")
        lines.append("| Название | Сумма | Тип привязки |")
        lines.append("|----------|-------|--------------|")
        for name, data in cv.items():
            lines.append(f"| {name} | {data.get('value', 0):,.0f} ₽ | {data.get('type', '')} |".replace(",", " "))
        lines.append("")

    lines.append("## Помесячные результаты (расчёт калькулятора)")
    lines.append("")
    lines.append(
        "| Месяц | Revenue | Fixed | Variable | Total costs | CF | Cumulative CF |"
    )
    lines.append("|-------|---------|-------|----------|-------------|-----|---------------|")
    for cf in cash_flow_results:
        lines.append(
            f"| {cf['month']} | {_fmt_rub(cf['revenue'])} | {_fmt_rub(cf['fixed_costs'])} | "
            f"{_fmt_rub(cf['variable_costs'])} | {_fmt_rub(cf['total_costs'])} | "
            f"{_fmt_rub(cf['cash_flow'])} | {_fmt_rub(cf['cumulative_cash_flow'])} |"
        )
    lines.append("")
    lines.append("### Детализация выручки по месяцам (ключевые поля)")
    lines.append("")
    lines.append("| Месяц | Ключевые поля (JSON-подобно) |")
    lines.append("|-------|------------------------------|")
    for r in revenue_results:
        m = r.get("month", "?")
        keys = {k: v for k, v in r.items() if k != "month"}
        lines.append(f"| {m} | `{json.dumps(keys, ensure_ascii=False)}` |")
    lines.append("")
    lines.append("## Фрагмент для вставки в БП_04 (черновик)")
    lines.append("")
    lines.append("Ниже — готовый текстовый блок; ассистент может адаптировать заголовки под структуру БП_04.")
    lines.append("")
    lines.append("### Допущения и границы модели")
    lines.append("")
    lines.append(
        f"- Горизонт расчёта в калькуляторе: **{num_months}** мес.; масштабирование на 24–36 мес. в БП_04 — отдельно."
    )
    lines.append(
        f"- Бизнес-модель в снимке: **{model_display_name}**; сценарий: **{scenario_label}**."
    )
    lines.append(
        "- Статус цифр: **〔синтез, проверить〕** — получены из интерактивного калькулятора, требуют согласования с командой и БП_01."
    )
    lines.append("")
    lines.append("### Краткий P&L (сумма за период)")
    lines.append("")
    lines.append(f"- Выручка: {_fmt_rub(total_rev)}")
    lines.append(f"- Затраты: {_fmt_rub(total_costs)}")
    lines.append(f"- Чистый денежный поток: {_fmt_rub(net_cf)}")
    lines.append("")
    lines.append("### Cash-flow и безубыточность")
    lines.append("")
    lines.append(f"- Breakeven (по кумулятивному CF в проекции калькулятора): **{be_text}**.")
    if min_rental_price != float("inf"):
        lines.append(f"- Ориентир минимальной цены аренды (Model B): **{_fmt_rub(min_rental_price)}**.")
    lines.append("")
    lines.append("### Следующие шаги для БП_04")
    lines.append("")
    lines.append("- Сверить параметры с БП_01 (A01–A04 и др.).")
    lines.append("- Добавить чувствительность и сценарии conservative/base/aggressive в едином документе.")
    lines.append("- Связать с Этапом 4 (коммерческая модель) и Этапом 7 (регуляторно-коммерческий трек).")
    lines.append("")

    # R&D фаза — добавляем раздел, если включена
    if rnd_results:
        total_rnd = sum(r["total_costs"] for r in rnd_results)
        lines.append("## R&D фаза (до старта продаж)")
        lines.append("")
        lines.append(
            f"- Длительность R&D: **{len(rnd_results)} мес.** — предшествует рыночной фазе."
        )
        lines.append(
            f"- Суммарные расходы R&D: **{_fmt_rub(total_rnd)}** — списываются из банка инвестиций."
        )
        lines.append(
            "- NPV рыночных месяцев дисконтируется с учётом сдвига: "
            f"рыночный месяц 1 = реальный месяц {len(rnd_results) + 1} от инвестиций."
        )
        lines.append("")
        lines.append("### Расходы R&D по месяцам")
        lines.append("")

        # Определяем все категории
        all_cats: List[str] = []
        for r in rnd_results:
            for cat in r.get("breakdown", {}).keys():
                if cat not in all_cats:
                    all_cats.append(cat)

        # Заголовок таблицы
        header = "| Период | " + " | ".join(all_cats) + " | Итого |"
        sep = "|--------|" + "--------|" * len(all_cats) + "--------|"
        lines.append(header)
        lines.append(sep)
        for r in rnd_results:
            row_parts = [f"R&D {r['month']}"]
            for cat in all_cats:
                row_parts.append(_fmt_rub(r["breakdown"].get(cat, 0.0)))
            row_parts.append(_fmt_rub(r["total_costs"]))
            lines.append("| " + " | ".join(row_parts) + " |")
        lines.append("")

    return "\n".join(lines)


def export_to_json(
    cash_flow_results: List[Dict],
    revenue_results: List[Dict],
    costs_results: List[Dict],
    unit_economics: Dict,
    all_params: Dict,
    filename: str = "reflex_calculator_results.json"
) -> str:
    """
    Экспорт результатов в JSON файл
    
    Returns:
        Путь к созданному файлу
    """
    try:
        data = {
            'parameters': all_params,
            'cash_flow': cash_flow_results,
            'revenue': revenue_results,
            'costs': costs_results,
            'unit_economics': unit_economics
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        return filename
    
    except Exception as e:
        return f"Ошибка при экспорте: {str(e)}"


def create_detailed_table(
    cash_flow_results: List[Dict],
    revenue_results: List[Dict],
    costs_results: List[Dict]
) -> pd.DataFrame:
    """
    Создание детальной таблицы расчетов для отображения
    
    Returns:
        DataFrame с детальными расчетами
    """
    data = []
    
    for i, cf in enumerate(cash_flow_results):
        revenue = revenue_results[i]
        costs = costs_results[i]
        
        row = {
            'Месяц': cf['month'],
            'Revenue (₽)': f"{cf['revenue']:,.0f}",
            'Fixed Costs (₽)': f"{cf['fixed_costs']:,.0f}",
            'Variable Costs (₽)': f"{cf['variable_costs']:,.0f}",
            'Total Costs (₽)': f"{cf['total_costs']:,.0f}",
            'Cash Flow (₽)': f"{cf['cash_flow']:,.0f}",
            'Cumulative CF (₽)': f"{cf['cumulative_cash_flow']:,.0f}"
        }
        
        data.append(row)
    
    return pd.DataFrame(data)
